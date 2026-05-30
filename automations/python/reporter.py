"""
MatchPro™ — Excel Report Generator
Produces a full .xlsx report per cycle.
"""
import os, sqlite3
from datetime import datetime
from config import REPORTS_DIR, DB_PATH

def generate_report(cycle_id: str, matches: list, summary: dict,
                    date_str: str, cycle_label: str) -> str:
    """
    Generate Excel report. Returns file path.
    Falls back to CSV if openpyxl not available.
    """
    os.makedirs(REPORTS_DIR, exist_ok=True)
    filename = f"MatchPro_Confirmations_{date_str}_{cycle_label}.xlsx"
    filepath = os.path.join(REPORTS_DIR, filename)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        _generate_xlsx(filepath, matches, summary, date_str, cycle_label)
    except ImportError:
        # Fallback to CSV
        filepath = filepath.replace(".xlsx", ".csv")
        _generate_csv(filepath, matches, summary)

    return filepath

def _generate_xlsx(filepath, matches, summary, date_str, cycle_label):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()

    # ── Sheet 1: Executive Summary ────────────────────────────────
    ws1 = wb.active
    ws1.title = "الملخص التنفيذي"

    ws1["A1"] = f"MatchPro™ | تقرير التأكيدات | {date_str} - {cycle_label}"
    ws1["A1"].font = Font(bold=True, size=14)
    ws1.merge_cells("A1:D1")

    headers = ["البيان", "العدد"]
    rows = [
        ["📨 إيميلات اتحللت", summary.get("emails_read", 0)],
        ["🔗 ماتشات اتعملت", summary.get("total_matches", 0)],
        ["✅ ماتشات مؤكدة", summary.get("confirmed", 0)],
        ["⚠️ محتاجة متابعة", summary.get("needs_followup", 0)],
        ["❌ مش مناسبة", summary.get("not_suitable", 0)],
        ["📈 متوسط الـ Score", f"{summary.get('avg_score', 0):.0f}%"],
    ]
    ws1.append([])
    ws1.append(headers)
    for row in rows:
        ws1.append(row)

    # ── Sheet 2: Confirmed Matches ────────────────────────────────
    ws2 = wb.create_sheet("✅ مؤكدة")
    ws2.append(["Match ID", "Score", "موقع العقار", "النوع", "السعر",
                "حالة البائع", "حالة المشتري"])
    for m in matches:
        if "Full Confirmed" in m.get("overall_status", ""):
            ws2.append([
                m.get("match_id", ""),
                f"{m.get('score', 0):.0f}%",
                m.get("location_ar", ""),
                m.get("prop_type", ""),
                f"{int(m.get('price', 0) or 0):,} EGP",
                m.get("seller_status", ""),
                m.get("buyer_status", ""),
            ])

    # ── Sheet 3: Needs Follow-up ──────────────────────────────────
    ws3 = wb.create_sheet("⚠️ تحتاج متابعة")
    ws3.append(["Match ID", "Score", "السبب", "الخطوة الجاية"])
    for m in matches:
        if "Needs Follow-up" in m.get("overall_status", ""):
            reason = _get_followup_reason(m)
            ws3.append([
                m.get("match_id", ""),
                f"{m.get('score', 0):.0f}%",
                reason,
                "متابعة الرد",
            ])

    # ── Sheet 4: Not Suitable ─────────────────────────────────────
    ws4 = wb.create_sheet("❌ غير مناسبة")
    ws4.append(["Match ID", "Score", "سبب الرفض"])
    for m in matches:
        if "Not Suitable" in m.get("overall_status", ""):
            ws4.append([
                m.get("match_id", ""),
                f"{m.get('score', 0):.0f}%",
                m.get("buyer_status", m.get("seller_status", "Low Score")),
            ])

    # ── Sheet 5: All Matches ──────────────────────────────────────
    ws5 = wb.create_sheet("كل الماتشات")
    ws5.append(["Match ID", "Score", "Location", "Type", "Price",
                "Seller", "Buyer", "Seller Status", "Buyer Status", "Overall"])
    for m in matches:
        ws5.append([
            m.get("match_id", ""),
            f"{m.get('score', 0):.0f}%",
            m.get("location_ar", ""),
            m.get("prop_type", ""),
            f"{int(m.get('price', 0) or 0):,}",
            m.get("seller_name", ""),
            m.get("buyer_name", ""),
            m.get("seller_status", ""),
            m.get("buyer_status", ""),
            m.get("overall_status", ""),
        ])

    wb.save(filepath)
    print(f"✅ Report saved: {filepath}")

def _generate_csv(filepath, matches, summary):
    import csv
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["Match ID", "Score", "Location", "Type", "Price",
                         "Seller Status", "Buyer Status", "Overall"])
        for m in matches:
            writer.writerow([
                m.get("match_id"), m.get("score"),
                m.get("location_ar"), m.get("prop_type"),
                m.get("price"), m.get("seller_status"),
                m.get("buyer_status"), m.get("overall_status"),
            ])
    print(f"✅ CSV report saved: {filepath}")

def _get_followup_reason(m):
    ss = m.get("seller_status", "")
    bs = m.get("buyer_status", "")
    if ss == "No Response": return "البائع ما ردش"
    if bs == "No Response": return "المشتري ما ردش"
    if ss == "Pending": return "في انتظار رد البائع"
    if bs == "Pending": return "في انتظار رد المشتري"
    return "محتاج مراجعة"
