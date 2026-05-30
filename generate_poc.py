#!/usr/bin/env python3
"""
MatchPro™ POC PDF Generator + Investor Excel Updater
Crystal Power Investments | May 7, 2026
"""

import openpyxl
import math
import shutil
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm, mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    HRFlowable, PageBreak, KeepTogether
)
from reportlab.pdfgen import canvas as pdfcanvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus.flowables import Flowable
import os

# ─────────────────────────────────────────────────────────────────────────────
# COLORS
# ─────────────────────────────────────────────────────────────────────────────
NAVY = colors.HexColor('#0B2545')
GOLD = colors.HexColor('#C9A961')
WHITE = colors.white
LIGHT_GREY = colors.HexColor('#F5F5F5')
ORANGE = colors.HexColor('#E07B39')
LIGHT_GOLD_BG = colors.HexColor('#FDF6E3')
WATERMARK_COLOR = colors.HexColor('#E8E8E8')

# ─────────────────────────────────────────────────────────────────────────────
# PATHS
# ─────────────────────────────────────────────────────────────────────────────
MATCHES_FILE = '/home/work/.openclaw/media/inbound/CrystalPower-MatchPro-2026-05-07---520c25a7-c206-4c41-a3a1-f3c168564b6d.xlsx'
INVESTOR_FILE = '/home/work/.openclaw/media/inbound/Investor_report_for_MatchPro_supply_and_demand_num-Genspark_---ca958824-3228-44d6-8de6-06885710e889.xlsx'
PDF_OUTPUT = '/home/work/.openclaw/workspace/matchpro/MatchPro_POC_Top100_Matches.pdf'
EXCEL_OUTPUT = '/home/work/.openclaw/workspace/matchpro/MatchPro_Investor_Report_COMPLETE.xlsx'

os.makedirs('/home/work/.openclaw/workspace/matchpro', exist_ok=True)


# ─────────────────────────────────────────────────────────────────────────────
# STEP 1: LOAD & FILTER MATCHES
# ─────────────────────────────────────────────────────────────────────────────
def safe_float(v, default=0.0):
    try:
        return float(str(v).replace('%', '').strip())
    except:
        return default

def load_top100():
    wb = openpyxl.load_workbook(MATCHES_FILE)
    ws = wb['Matches']
    
    headers = [cell.value for cell in ws[4]]
    all_rows = []
    for row_idx in range(5, ws.max_row + 1):
        row = [cell.value for cell in ws[row_idx]]
        if any(v is not None for v in row):
            all_rows.append(row)
    
    print(f"Total rows loaded: {len(all_rows)}")
    
    valid = []
    skip_phone = 0
    skip_budget = 0
    
    for r in all_rows:
        score = safe_float(r[1])
        transaction = str(r[3]).upper().strip() if r[3] else ''
        seller_phone = str(r[5]).strip() if r[5] else ''
        buyer_phone = str(r[14]).strip() if r[14] else ''
        budget_max = safe_float(r[16]) if r[16] is not None else 0
        
        # Skip same phone
        if seller_phone and buyer_phone and seller_phone == buyer_phone:
            skip_phone += 1
            continue
        
        # Skip SALE with budget < 10000
        if transaction == 'SALE' and budget_max < 10000:
            skip_budget += 1
            continue
        
        valid.append((score, r))
    
    print(f"Skipped (same phone): {skip_phone}")
    print(f"Skipped (low budget SALE): {skip_budget}")
    print(f"Valid after filtering: {len(valid)}")
    
    # Sort descending by score
    valid.sort(key=lambda x: x[0], reverse=True)
    top100 = valid[:100]
    
    print(f"Top 100 score range: {top100[-1][0]:.1f}% — {top100[0][0]:.1f}%")
    return [r for _, r in top100], headers

TOP100_ROWS, HEADERS = load_top100()

# ─────────────────────────────────────────────────────────────────────────────
# COMPUTE STATS
# ─────────────────────────────────────────────────────────────────────────────
scores = [safe_float(r[1]) for r in TOP100_ROWS]
avg_score = sum(scores) / len(scores)
count_90plus = sum(1 for s in scores if s >= 90)
count_85_89 = sum(1 for s in scores if 85 <= s < 90)
count_80_84 = sum(1 for s in scores if 80 <= s < 85)
count_below_80 = sum(1 for s in scores if s < 80)

sale_count = sum(1 for r in TOP100_ROWS if str(r[3]).upper().strip() == 'SALE')
rent_count = sum(1 for r in TOP100_ROWS if str(r[3]).upper().strip() in ['RENT', 'RENTAL', 'LEASE'])

unique_locations = set()
for r in TOP100_ROWS:
    loc = str(r[8]).strip() if r[8] else ''
    if loc and loc.lower() != 'none':
        unique_locations.add(loc)

print(f"\nStats:")
print(f"  Avg score: {avg_score:.1f}%")
print(f"  >=90: {count_90plus}, 85-89: {count_85_89}, 80-84: {count_80_84}, <80: {count_below_80}")
print(f"  Sale: {sale_count}, Rent: {rent_count}")
print(f"  Unique locations: {len(unique_locations)}")

# ─────────────────────────────────────────────────────────────────────────────
# PDF CANVAS DECORATOR — WATERMARK + FOOTER + HEADER
# ─────────────────────────────────────────────────────────────────────────────
class DocTemplate(SimpleDocTemplate):
    def __init__(self, filename, **kwargs):
        super().__init__(filename, **kwargs)
        self.page_num = 0

def add_watermark_and_footer(canvas_obj, doc):
    """Add CONFIDENTIAL watermark and footer to every page."""
    width, height = A4
    
    # Save state
    canvas_obj.saveState()
    
    # Watermark
    canvas_obj.setFont('Helvetica-Bold', 55)
    canvas_obj.setFillColor(WATERMARK_COLOR)
    canvas_obj.setFillAlpha(0.25)
    canvas_obj.translate(width / 2, height / 2)
    canvas_obj.rotate(45)
    canvas_obj.drawCentredString(0, 0, 'CONFIDENTIAL')
    canvas_obj.rotate(-45)
    canvas_obj.translate(-width / 2, -height / 2)
    
    # Footer
    canvas_obj.setFillAlpha(1.0)
    canvas_obj.setFillColor(NAVY)
    canvas_obj.setFont('Helvetica', 7)
    footer_text = f'Confidential — Crystal Power Investments © 2026 | MatchPro™ POC'
    canvas_obj.drawCentredString(width / 2, 12 * mm, footer_text)
    canvas_obj.setStrokeColor(GOLD)
    canvas_obj.setLineWidth(0.5)
    canvas_obj.line(20 * mm, 15 * mm, width - 20 * mm, 15 * mm)
    
    canvas_obj.restoreState()


# ─────────────────────────────────────────────────────────────────────────────
# HELPER: Format value for display
# ─────────────────────────────────────────────────────────────────────────────
def fmt_price(val):
    if val is None:
        return 'N/A'
    try:
        n = float(str(val).replace(',', ''))
        if n >= 1_000_000:
            return f'{n/1_000_000:.1f}M'
        elif n >= 1_000:
            return f'{n/1_000:.0f}K'
        return str(int(n))
    except:
        return str(val) if val else 'N/A'

def fmt_str(val, maxlen=20):
    if val is None:
        return 'N/A'
    s = str(val).strip()
    if not s or s.lower() == 'none':
        return 'N/A'
    # Remove emoji from score text
    if len(s) > maxlen:
        s = s[:maxlen-1] + '…'
    return s

def fmt_name(val, maxlen=15):
    return fmt_str(val, maxlen)

def score_cell_color(score):
    if score >= 90:
        return GOLD
    elif score >= 85:
        return ORANGE
    else:
        return None  # default


# ─────────────────────────────────────────────────────────────────────────────
# BUILD PDF
# ─────────────────────────────────────────────────────────────────────────────
def build_pdf():
    doc = SimpleDocTemplate(
        PDF_OUTPUT,
        pagesize=A4,
        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=2 * cm,
        bottomMargin=2.5 * cm,
    )
    
    story = []
    width, height = A4
    
    # ── PAGE 1: COVER ──────────────────────────────────────────────────────────
    # Full navy background via a custom flowable
    class CoverPage(Flowable):
        def __init__(self, w, h):
            super().__init__()
            self.w = w
            self.h = h
        
        def draw(self):
            c = self.canv
            cw, ch = A4
            
            # Full navy background
            c.setFillColor(NAVY)
            c.rect(-1.5*cm, -2.5*cm, cw + 3*cm, ch + 4.5*cm, fill=1, stroke=0)
            
            # Gold accent line at top
            c.setStrokeColor(GOLD)
            c.setLineWidth(4)
            c.line(-1.5*cm, ch - 2*cm + 2.5*cm, cw - 1.5*cm, ch - 2*cm + 2.5*cm)
            
            # MatchPro™ title
            c.setFillColor(GOLD)
            c.setFont('Helvetica-Bold', 52)
            c.drawCentredString(cw/2 - 1.5*cm, ch - 5*cm, 'MatchPro\u2122')
            
            # Subtitle
            c.setFillColor(WHITE)
            c.setFont('Helvetica-Bold', 20)
            c.drawCentredString(cw/2 - 1.5*cm, ch - 7*cm, 'Proof of Concept \u2014 Top 100 Matches')
            
            # Arabic subtitle
            c.setFont('Helvetica', 14)
            c.drawCentredString(cw/2 - 1.5*cm, ch - 8.2*cm, '\u0625\u062b\u0628\u0627\u062a \u0627\u0644\u062c\u062f\u0648\u0649 \u2014 \u0623\u0641\u0636\u0644 100 \u0645\u0637\u0627\u0628\u0642\u0629')
            
            # Gold rule
            c.setStrokeColor(GOLD)
            c.setLineWidth(2)
            c.line(3*cm, ch - 9.5*cm, cw - 4.5*cm, ch - 9.5*cm)
            
            # Company + date
            c.setFillColor(WHITE)
            c.setFont('Helvetica-Bold', 13)
            c.drawCentredString(cw/2 - 1.5*cm, ch - 10.8*cm, 'Crystal Power Investments  |  May 7, 2026')
            
            # Tagline
            c.setFont('Helvetica-Oblique', 11)
            c.setFillColor(GOLD)
            c.drawCentredString(cw/2 - 1.5*cm, ch - 12.5*cm,
                'The market is talking \u2014 are you listening?')
            c.setFillColor(WHITE)
            c.drawCentredString(cw/2 - 1.5*cm, ch - 13.5*cm,
                '\u0627\u0644\u0633\u0648\u0642 \u064a\u062a\u0643\u0644\u0645\u060c \u0647\u0644 \u0623\u0646\u062a \u062a\u064f\u0646\u0635\u062a\u061f')
            
            # Stats teaser box
            c.setFillColor(colors.HexColor('#112E5C'))
            c.roundRect(1*cm, 4*cm, cw - 5*cm, 6*cm, 8, fill=1, stroke=0)
            
            c.setStrokeColor(GOLD)
            c.setLineWidth(1.5)
            c.roundRect(1*cm, 4*cm, cw - 5*cm, 6*cm, 8, fill=0, stroke=1)
            
            stats = [
                ('500', 'Total Matches Analyzed'),
                ('100', 'Top Matches Selected'),
                (f'{avg_score:.1f}%', 'Average Score'),
                (f'{len(unique_locations)}', 'Cairo Areas Covered'),
            ]
            col_w = (cw - 5*cm) / 4
            for i, (val, label) in enumerate(stats):
                x = 1*cm + i * col_w + col_w/2
                c.setFillColor(GOLD)
                c.setFont('Helvetica-Bold', 22)
                c.drawCentredString(x, 9*cm, val)
                c.setFillColor(WHITE)
                c.setFont('Helvetica', 9)
                c.drawCentredString(x, 8*cm, label)
            
            # Confidential notice
            c.setFillColor(colors.HexColor('#888888'))
            c.setFont('Helvetica-Oblique', 8)
            c.drawCentredString(cw/2 - 1.5*cm, 1.5*cm, 
                'CONFIDENTIAL — For authorized recipients of Crystal Power Investments only')
            
            # Watermark
            c.saveState()
            c.setFont('Helvetica-Bold', 55)
            c.setFillColor(colors.HexColor('#1A3A66'))
            c.setFillAlpha(0.15)
            c.translate(cw/2 - 1.5*cm, ch/2)
            c.rotate(45)
            c.drawCentredString(0, 0, 'CONFIDENTIAL')
            c.restoreState()
        
        def wrap(self, availWidth, availHeight):
            return (availWidth, availHeight)
    
    story.append(CoverPage(width, height))
    story.append(PageBreak())
    
    # ── PAGE 2: EXECUTIVE SUMMARY ──────────────────────────────────────────────
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'PageTitle',
        fontName='Helvetica-Bold',
        fontSize=22,
        textColor=NAVY,
        spaceAfter=6,
        spaceBefore=0,
        leading=26,
    )
    
    subtitle_style = ParagraphStyle(
        'Subtitle',
        fontName='Helvetica',
        fontSize=11,
        textColor=colors.HexColor('#555555'),
        spaceAfter=14,
    )
    
    body_style = ParagraphStyle(
        'Body',
        fontName='Helvetica',
        fontSize=10,
        textColor=NAVY,
        spaceAfter=8,
        leading=15,
    )
    
    label_style = ParagraphStyle(
        'Label',
        fontName='Helvetica-Bold',
        fontSize=9,
        textColor=NAVY,
        alignment=TA_CENTER,
    )
    
    value_style = ParagraphStyle(
        'Value',
        fontName='Helvetica-Bold',
        fontSize=20,
        textColor=GOLD,
        alignment=TA_CENTER,
        leading=24,
    )
    
    story.append(Paragraph('POC at a Glance', title_style))
    story.append(HRFlowable(width='100%', thickness=2, color=GOLD, spaceAfter=14))
    
    # Stat boxes — 4 per row
    stat_data_1 = [
        [
            Paragraph('500', value_style),
            Paragraph('100', value_style),
            Paragraph(f'{avg_score:.1f}%', value_style),
            Paragraph(str(count_90plus), value_style),
        ],
        [
            Paragraph('Total Matches<br/>Analyzed', label_style),
            Paragraph('Top 100<br/>Selected', label_style),
            Paragraph('Avg Score<br/>(Top 100)', label_style),
            Paragraph('Score<br/>≥90%', label_style),
        ],
    ]
    
    stat_data_2 = [
        [
            Paragraph(str(count_85_89), value_style),
            Paragraph(str(sale_count), value_style),
            Paragraph(str(rent_count), value_style),
            Paragraph(str(len(unique_locations)), value_style),
        ],
        [
            Paragraph('Score<br/>85–89%', label_style),
            Paragraph('Sale<br/>Matches', label_style),
            Paragraph('Rent<br/>Matches', label_style),
            Paragraph('Unique<br/>Locations', label_style),
        ],
    ]
    
    box_style = TableStyle([
        ('BOX', (0, 0), (-1, -1), 1.5, NAVY),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('BACKGROUND', (0, 0), (-1, -1), LIGHT_GREY),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
    ])
    
    col_w = (doc.width) / 4
    
    t1 = Table(stat_data_1, colWidths=[col_w] * 4)
    t1.setStyle(box_style)
    story.append(t1)
    story.append(Spacer(1, 6))
    
    t2 = Table(stat_data_2, colWidths=[col_w] * 4)
    t2.setStyle(box_style)
    story.append(t2)
    story.append(Spacer(1, 14))
    
    # Key insight box
    insight_style = ParagraphStyle(
        'Insight',
        fontName='Helvetica',
        fontSize=10,
        textColor=NAVY,
        leading=15,
        leftIndent=12,
        rightIndent=8,
    )
    
    insight_text = (
        "These 100 matches represent the highest-confidence supply-demand pairings identified by "
        "MatchPro\u2122 from <b>500 total matches</b> generated from <b>50+ WhatsApp broker groups</b>. "
        "Each match has been scored across 3 dimensions: <b>Location</b>, <b>Price</b>, and <b>Specifications</b>."
    )
    
    insight_table = Table(
        [[Paragraph(insight_text, insight_style)]],
        colWidths=[doc.width]
    )
    insight_table.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 0, WHITE),
        ('LINEBEFORE', (0, 0), (0, -1), 4, GOLD),
        ('BACKGROUND', (0, 0), (-1, -1), LIGHT_GOLD_BG),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('LEFTPADDING', (0, 0), (-1, -1), 14),
        ('RIGHTPADDING', (0, 0), (-1, -1), 10),
    ]))
    story.append(insight_table)
    story.append(Spacer(1, 16))
    
    # Score breakdown table
    story.append(Paragraph('Score Distribution — Top 100', ParagraphStyle(
        'SectionHead', fontName='Helvetica-Bold', fontSize=13, textColor=NAVY, spaceAfter=8
    )))
    
    breakdown_ranges = [
        ('≥90%', count_90plus),
        ('85–89%', count_85_89),
        ('80–84%', count_80_84),
        ('<80%', count_below_80),
    ]
    breakdown_data = [
        [Paragraph('<b>Score Range</b>', label_style),
         Paragraph('<b>Count</b>', label_style),
         Paragraph('<b>% of Top 100</b>', label_style)]
    ]
    for rng, cnt in breakdown_ranges:
        if cnt > 0:
            breakdown_data.append([
                Paragraph(rng, ParagraphStyle('BD', fontName='Helvetica', fontSize=10, textColor=NAVY, alignment=TA_CENTER)),
                Paragraph(str(cnt), ParagraphStyle('BD', fontName='Helvetica-Bold', fontSize=10, textColor=GOLD, alignment=TA_CENTER)),
                Paragraph(f'{cnt}%', ParagraphStyle('BD', fontName='Helvetica', fontSize=10, textColor=NAVY, alignment=TA_CENTER)),
            ])
    
    bt = Table(breakdown_data, colWidths=[doc.width/3]*3)
    bt.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), NAVY),
        ('TEXTCOLOR', (0, 0), (-1, 0), WHITE),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('BOX', (0, 0), (-1, -1), 1, NAVY),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [WHITE, LIGHT_GREY]),
    ]))
    story.append(bt)
    story.append(PageBreak())
    
    # ── PAGE 3: PLATFORM STATS ──────────────────────────────────────────────────
    story.append(Paragraph('MatchPro\u2122 Platform — Real Numbers', title_style))
    story.append(HRFlowable(width='100%', thickness=2, color=GOLD, spaceAfter=14))
    
    platform_stats = [
        ('5,569', 'Broker messages processed', 'Supply signals ingested from 50+ WhatsApp broker groups'),
        ('4,000', 'Supply listings extracted', 'Active inventory tracked by the engine'),
        ('7,626', 'Demand requests extracted', 'Buyer intent signals — exceeds supply by 91%'),
        ('56,566', 'Total AI matches generated', 'Cross-pairs computed by the semantic engine'),
        ('1,438', 'High-confidence matches ≥90%', 'The Very Hot Leads — actionable right now'),
        ('40+', 'Greater Cairo areas covered', 'Madinaty, NC, Sheikh Zayed, NAC, North Coast, etc.'),
        ('$97.5M', 'Total pipeline value', 'EGP 4.87B aggregated buyer budgets ÷ EGP 50/USD'),
        ('91%', 'Demand surplus over supply', '7,626 buyers chasing 4,000 listings'),
    ]
    
    pstat_style_val = ParagraphStyle('PSVal', fontName='Helvetica-Bold', fontSize=17, textColor=GOLD)
    pstat_style_key = ParagraphStyle('PSKey', fontName='Helvetica-Bold', fontSize=10, textColor=NAVY)
    pstat_style_desc = ParagraphStyle('PSDesc', fontName='Helvetica', fontSize=9, textColor=colors.HexColor('#555555'), leading=12)
    
    pstat_data = []
    for val, key, desc in platform_stats:
        pstat_data.append([
            Paragraph(val, pstat_style_val),
            Paragraph(f'{key}<br/><font size="9" color="#555555">{desc}</font>', pstat_style_key),
        ])
    
    pt = Table(pstat_data, colWidths=[3*cm, doc.width - 3*cm])
    pt.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#DDDDDD')),
        ('BOX', (0, 0), (-1, -1), 1, NAVY),
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [WHITE, LIGHT_GREY]),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (0, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(pt)
    story.append(Spacer(1, 16))
    
    # Why this matters box
    why_text = (
        "<b>Why This Matters:</b> 7,626 buyers chasing 4,000 listings — a <b>91% demand surplus</b>. "
        "MatchPro\u2122 identifies WHERE the asymmetry is highest, in real time, every 12 hours. "
        "No other tool in the Egyptian market does this at scale."
    )
    why_table = Table(
        [[Paragraph(why_text, ParagraphStyle('Why', fontName='Helvetica', fontSize=10, textColor=NAVY, leading=15))]],
        colWidths=[doc.width]
    )
    why_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), LIGHT_GOLD_BG),
        ('LINEBEFORE', (0, 0), (0, -1), 4, GOLD),
        ('BOX', (0, 0), (-1, -1), 1, GOLD),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('LEFTPADDING', (0, 0), (-1, -1), 14),
        ('RIGHTPADDING', (0, 0), (-1, -1), 10),
    ]))
    story.append(why_table)
    story.append(PageBreak())
    
    # ── PAGES 4-13: MATCH TABLES (10 per page) ─────────────────────────────────
    match_col_style = ParagraphStyle('MCStyle', fontName='Helvetica', fontSize=8, textColor=NAVY, leading=11)
    match_hdr_style = ParagraphStyle('MHStyle', fontName='Helvetica-Bold', fontSize=8, textColor=WHITE, leading=11, alignment=TA_CENTER)
    match_score_90 = ParagraphStyle('MS90', fontName='Helvetica-Bold', fontSize=9, textColor=NAVY, alignment=TA_CENTER)
    match_score_85 = ParagraphStyle('MS85', fontName='Helvetica-Bold', fontSize=9, textColor=WHITE, alignment=TA_CENTER)
    match_score_norm = ParagraphStyle('MSN', fontName='Helvetica', fontSize=9, textColor=NAVY, alignment=TA_CENTER)
    
    col_widths = [
        0.7*cm,   # #
        1.3*cm,   # Score
        1.5*cm,   # Type
        1.5*cm,   # Txn
        3.0*cm,   # Location
        2.5*cm,   # Price EGP
        3.0*cm,   # Seller
        3.0*cm,   # Buyer
        2.5*cm,   # Budget EGP
    ]
    # Normalize to doc width
    total_cw = sum(col_widths)
    scale = doc.width / total_cw
    col_widths = [w * scale for w in col_widths]
    
    PAGES = [TOP100_ROWS[i:i+10] for i in range(0, 100, 10)]
    
    for page_idx, page_rows in enumerate(PAGES):
        start_num = page_idx * 10 + 1
        end_num = start_num + len(page_rows) - 1
        
        page_title = f'MatchPro\u2122 Top 100 \u2014 Matches #{start_num} to #{end_num}'
        story.append(Paragraph(page_title, title_style))
        story.append(HRFlowable(width='100%', thickness=1.5, color=GOLD, spaceAfter=8))
        
        # Build table rows
        table_data = [[
            Paragraph('#', match_hdr_style),
            Paragraph('Score%', match_hdr_style),
            Paragraph('Type', match_hdr_style),
            Paragraph('Transaction', match_hdr_style),
            Paragraph('Location', match_hdr_style),
            Paragraph('Price (EGP)', match_hdr_style),
            Paragraph('Seller', match_hdr_style),
            Paragraph('Buyer', match_hdr_style),
            Paragraph('Budget (EGP)', match_hdr_style),
        ]]
        
        row_colors = []
        score_cell_colors = {}
        
        for i, row in enumerate(page_rows):
            match_num = start_num + i
            score = safe_float(row[1])
            
            # Score style
            if score >= 90:
                s_style = match_score_90
                score_cell_colors[(i+1, 1)] = GOLD
            elif score >= 85:
                s_style = match_score_85
                score_cell_colors[(i+1, 1)] = ORANGE
            else:
                s_style = match_score_norm
            
            transaction = str(row[3]).upper() if row[3] else 'N/A'
            prop_type = fmt_str(row[7], 12)
            location = fmt_str(row[8], 18)
            price = fmt_price(row[9])
            seller = fmt_name(row[4], 14)
            buyer = fmt_name(row[13], 14)
            budget = fmt_price(row[16])
            
            bg = WHITE if i % 2 == 0 else LIGHT_GREY
            row_colors.append(bg)
            
            table_data.append([
                Paragraph(str(match_num), ParagraphStyle('N', fontName='Helvetica-Bold', fontSize=9, textColor=NAVY, alignment=TA_CENTER)),
                Paragraph(f'{score:.0f}%', s_style),
                Paragraph(prop_type, match_col_style),
                Paragraph(transaction, ParagraphStyle('T', fontName='Helvetica', fontSize=8, textColor=NAVY, alignment=TA_CENTER)),
                Paragraph(location, match_col_style),
                Paragraph(price, ParagraphStyle('P', fontName='Helvetica', fontSize=8, textColor=NAVY, alignment=TA_RIGHT)),
                Paragraph(seller, match_col_style),
                Paragraph(buyer, match_col_style),
                Paragraph(budget, ParagraphStyle('B', fontName='Helvetica', fontSize=8, textColor=NAVY, alignment=TA_RIGHT)),
            ])
        
        mt = Table(table_data, colWidths=col_widths, repeatRows=1)
        
        ts = TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), NAVY),
            ('TEXTCOLOR', (0, 0), (-1, 0), WHITE),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ('BOX', (0, 0), (-1, -1), 1, NAVY),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ])
        
        # Row backgrounds
        for i, bg in enumerate(row_colors):
            ts.add('BACKGROUND', (0, i+1), (-1, i+1), bg)
        
        # Score cell colors
        for (row_i, col_i), bg_color in score_cell_colors.items():
            ts.add('BACKGROUND', (col_i, row_i), (col_i, row_i), bg_color)
        
        mt.setStyle(ts)
        story.append(mt)
        story.append(PageBreak())
    
    # ── LAST PAGE: METHODOLOGY ─────────────────────────────────────────────────
    story.append(Paragraph('Methodology & Transparency', title_style))
    story.append(HRFlowable(width='100%', thickness=2, color=GOLD, spaceAfter=16))
    
    meth_style = ParagraphStyle(
        'Meth', fontName='Helvetica', fontSize=10, textColor=NAVY, leading=15, spaceAfter=10
    )
    bullet_style = ParagraphStyle(
        'Bullet', fontName='Helvetica', fontSize=10, textColor=NAVY, leading=15,
        leftIndent=20, spaceAfter=8, bulletIndent=8
    )
    
    story.append(Paragraph('<b>How MatchPro\u2122 Works</b>', ParagraphStyle(
        'H2', fontName='Helvetica-Bold', fontSize=13, textColor=NAVY, spaceAfter=10
    )))
    
    bullets = [
        "Messages from <b>50+ broker WhatsApp groups</b> are processed every <b>12 hours</b>",
        "Each supply listing is paired against every demand request across <b>3 scoring dimensions</b>: "
        "<b>Location</b> (area match), <b>Price</b> (budget alignment), <b>Specifications</b> (type, size, bedrooms)",
        "Only matches scoring <b>≥70% across all three dimensions</b> are surfaced — ensuring every match "
        "has real commercial potential",
    ]
    
    for b in bullets:
        story.append(Paragraph(f'\u2022 {b}', bullet_style))
    
    story.append(Spacer(1, 12))
    
    # Score legend
    story.append(Paragraph('<b>Score Interpretation</b>', ParagraphStyle(
        'H2', fontName='Helvetica-Bold', fontSize=13, textColor=NAVY, spaceAfter=10
    )))
    
    legend_data = [
        [Paragraph('<b>Score Range</b>', match_hdr_style),
         Paragraph('<b>Color</b>', match_hdr_style),
         Paragraph('<b>Meaning</b>', match_hdr_style)],
        [Paragraph('≥90%', match_score_90), Paragraph('Gold', match_col_style), 
         Paragraph('Highest confidence — near-certain match', match_col_style)],
        [Paragraph('85–89%', match_score_85), Paragraph('Orange', match_col_style),
         Paragraph('High confidence — strong commercial alignment', match_col_style)],
        [Paragraph('70–84%', match_score_norm), Paragraph('White', match_col_style),
         Paragraph('Good confidence — merits broker follow-up', match_col_style)],
    ]
    
    lt = Table(legend_data, colWidths=[3*cm, 2.5*cm, doc.width - 5.5*cm])
    lt.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), NAVY),
        ('BACKGROUND', (0, 1), (-1, 1), LIGHT_GOLD_BG),
        ('BACKGROUND', (0, 2), (0, 2), ORANGE),
        ('BACKGROUND', (0, 3), (-1, 3), LIGHT_GREY),
        ('BOX', (0, 0), (-1, -1), 1, NAVY),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (1, -1), 'CENTER'),
    ]))
    story.append(lt)
    story.append(Spacer(1, 20))
    
    # Disclaimer
    disclaimer_text = (
        "This document is <b>confidential</b> and intended solely for authorized recipients of "
        "Crystal Power Investments. The matches presented are generated by automated analysis "
        "and should be verified by qualified real estate professionals before commercial action."
    )
    
    disc_table = Table(
        [[Paragraph(disclaimer_text, ParagraphStyle('Disc', fontName='Helvetica-Oblique', fontSize=9, textColor=colors.HexColor('#555555'), leading=13))]],
        colWidths=[doc.width]
    )
    disc_table.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#CCCCCC')),
        ('BACKGROUND', (0, 0), (-1, -1), LIGHT_GREY),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('LEFTPADDING', (0, 0), (-1, -1), 12),
        ('RIGHTPADDING', (0, 0), (-1, -1), 12),
    ]))
    story.append(disc_table)
    story.append(Spacer(1, 12))
    
    # Footer attribution
    story.append(Paragraph(
        'Crystal Power Investments | MatchPro\u2122 Intelligence Engine | May 2026',
        ParagraphStyle('Footer', fontName='Helvetica-Bold', fontSize=9, textColor=NAVY, alignment=TA_CENTER)
    ))
    
    # Build PDF
    print(f"\nBuilding PDF: {PDF_OUTPUT}")
    doc.build(story, onFirstPage=add_watermark_and_footer, onLaterPages=add_watermark_and_footer)
    print(f"PDF generated successfully!")
    return True


# ─────────────────────────────────────────────────────────────────────────────
# STEP 2: EXCEL — Copy investor file & add Top100 sheet
# ─────────────────────────────────────────────────────────────────────────────
def build_excel():
    print(f"\nCopying investor Excel to: {EXCEL_OUTPUT}")
    shutil.copy2(INVESTOR_FILE, EXCEL_OUTPUT)
    
    wb = openpyxl.load_workbook(EXCEL_OUTPUT)
    print(f"Sheets found: {wb.sheetnames}")
    
    # ── Add/replace Top100_Matches sheet ──────────────────────────────────────
    if 'Top100_Matches' in wb.sheetnames:
        del wb['Top100_Matches']
    
    ws_top = wb.create_sheet('Top100_Matches')
    
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    
    NAVY_HEX = '0B2545'
    GOLD_HEX = 'C9A961'
    ORANGE_HEX = 'E07B39'
    LIGHT_GREY_HEX = 'F5F5F5'
    
    navy_fill = PatternFill('solid', fgColor=NAVY_HEX)
    gold_fill = PatternFill('solid', fgColor=GOLD_HEX)
    orange_fill = PatternFill('solid', fgColor=ORANGE_HEX)
    grey_fill = PatternFill('solid', fgColor=LIGHT_GREY_HEX)
    
    thin = Side(style='thin', color='CCCCCC')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    # Title rows
    ws_top.merge_cells('A1:Z1')
    ws_top['A1'] = 'MatchPro™ — Top 100 Matches | Proof of Concept | May 7, 2026'
    ws_top['A1'].font = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
    ws_top['A1'].fill = navy_fill
    ws_top['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_top.row_dimensions[1].height = 28
    
    ws_top.merge_cells('A2:Z2')
    ws_top['A2'] = f'Crystal Power Investments | {count_90plus} matches ≥90% | Avg Score: {avg_score:.1f}% | {sale_count} Sale | {rent_count} Rent | {len(unique_locations)} Locations'
    ws_top['A2'].font = Font(name='Calibri', bold=True, size=10, color=GOLD_HEX)
    ws_top['A2'].fill = navy_fill
    ws_top['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws_top.row_dimensions[2].height = 18
    
    # Headers (row 3)
    display_headers = [
        '#', 'Match ID', 'Score %', 'Status', 'Transaction',
        'Seller Name', 'Seller Phone', 'Seller Group',
        'Property Type', 'Location', 'Price (EGP)', 'Bedrooms', 'Size (m²)',
        'Buyer Name', 'Buyer Phone', 'Buyer Group',
        'Budget Max (EGP)', 'Buyer Bedrooms', 'Demand Location',
        'Location Score %', 'Price Score %', 'Specs Score %',
        'Conversion Stage', 'Last Contacted', 'Match Date',
    ]
    
    for col_idx, hdr in enumerate(display_headers, 1):
        cell = ws_top.cell(row=3, column=col_idx, value=hdr)
        cell.font = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws_top.row_dimensions[3].height = 32
    
    # Data rows
    col_mapping = [
        None,  # # (rank)
        0,   # Match ID
        1,   # Score %
        2,   # Status
        3,   # Transaction
        4,   # Seller Name
        5,   # Seller Phone
        6,   # Seller Group
        7,   # Property Type
        8,   # Location
        9,   # Price (EGP)
        10,  # Bedrooms
        11,  # Size (m²)
        13,  # Buyer Name
        14,  # Buyer Phone
        15,  # Buyer Group
        16,  # Budget Max (EGP)
        17,  # Buyer Bedrooms
        18,  # Demand Location
        20,  # Location Score %
        21,  # Price Score %
        22,  # Specs Score %
        23,  # Conversion Stage
        24,  # Last Contacted
        25,  # Match Date
    ]
    
    for rank, row in enumerate(TOP100_ROWS, 1):
        excel_row = rank + 3
        score = safe_float(row[1])
        
        # Row background
        if rank % 2 == 0:
            row_fill = grey_fill
        else:
            row_fill = PatternFill('solid', fgColor='FFFFFF')
        
        for col_idx, src_col in enumerate(col_mapping, 1):
            cell = ws_top.cell(row=excel_row, column=col_idx)
            
            if src_col is None:
                # Rank column
                cell.value = rank
                cell.font = Font(name='Calibri', bold=True, size=10, color=NAVY_HEX)
                cell.fill = row_fill
            else:
                val = row[src_col]
                if val is not None and str(val).lower() != 'none':
                    # Try to keep numbers as numbers
                    try:
                        cell.value = float(val) if '.' in str(val) else int(val)
                    except:
                        cell.value = str(val)
                else:
                    cell.value = None
                
                # Score column highlight
                if col_idx == 3:  # Score %
                    if score >= 90:
                        cell.fill = gold_fill
                        cell.font = Font(name='Calibri', bold=True, size=10, color=NAVY_HEX)
                    elif score >= 85:
                        cell.fill = orange_fill
                        cell.font = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
                    else:
                        cell.fill = row_fill
                        cell.font = Font(name='Calibri', size=10, color=NAVY_HEX)
                else:
                    cell.fill = row_fill
                    cell.font = Font(name='Calibri', size=10, color=NAVY_HEX)
            
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=False)
        
        ws_top.row_dimensions[excel_row].height = 16
    
    # Column widths
    col_widths_xlsx = [5, 12, 9, 10, 12, 18, 14, 20, 14, 18, 14, 10, 10, 18, 14, 20, 14, 12, 18, 12, 12, 12, 16, 14, 12]
    for i, w in enumerate(col_widths_xlsx, 1):
        ws_top.column_dimensions[get_column_letter(i)].width = w
    
    # Freeze panes
    ws_top.freeze_panes = 'A4'
    
    # ── Update Transparency_Messages sheet — add summary section ──────────────
    if 'Transparency_Messages' in wb.sheetnames:
        ws_trans = wb['Transparency_Messages']
        # Find next empty row
        next_row = ws_trans.max_row + 2
        ws_trans.merge_cells(f'A{next_row}:G{next_row}')
        ws_trans[f'A{next_row}'] = '─── Top 100 Match Summary (May 7, 2026) ───'
        ws_trans[f'A{next_row}'].font = Font(bold=True, size=11, color=NAVY_HEX)
        
        summary_rows = [
            ('Total matches analyzed', 500),
            ('Valid after filtering', len([r for r in TOP100_ROWS])),
            ('Avg score (top 100)', f'{avg_score:.1f}%'),
            ('Matches ≥90%', count_90plus),
            ('Matches 85-89%', count_85_89),
            ('Sale matches', sale_count),
            ('Rent matches', rent_count),
            ('Unique locations', len(unique_locations)),
        ]
        for label, val in summary_rows:
            next_row += 1
            ws_trans[f'A{next_row}'] = label
            ws_trans[f'B{next_row}'] = val
    
    wb.save(EXCEL_OUTPUT)
    print(f"Excel saved: {EXCEL_OUTPUT}")
    return True


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    print("=" * 60)
    print("MatchPro™ POC Generator")
    print("=" * 60)
    
    pdf_ok = build_pdf()
    excel_ok = build_excel()
    
    print("\n" + "=" * 60)
    print("OUTPUT FILES:")
    if pdf_ok:
        import os
        size = os.path.getsize(PDF_OUTPUT)
        print(f"  PDF:   {PDF_OUTPUT}")
        print(f"         Size: {size/1024:.1f} KB")
    if excel_ok:
        import os
        size = os.path.getsize(EXCEL_OUTPUT)
        print(f"  Excel: {EXCEL_OUTPUT}")
        print(f"         Size: {size/1024:.1f} KB")
    print("=" * 60)
