#!/usr/bin/env python3
"""
AUTO PROPERTY MATCHER — Crystal Power Investments
Runs on any new MatchPro Excel file → generates matched Excel per property → sends via WhatsApp
Zero external credits. Uses Green API.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import sys, os, json, subprocess
from datetime import datetime

INSTANCE = "7105409203"
TOKEN = "0e7ca429980f4331ae5fee4360c955a9db2d6fe3ca6545a4b3"
MOMEN = "201066505665@c.us"
WEBROOT = "/var/www/html"
BASE_URL = "https://lkdsbjzk.gensparkclaw.com"

# CPI ASSETS (add/update here as portfolio changes)
CPI_ASSETS = {
    "B11":      {"name": "A-308605 B11 3BR 133m²", "price": "5,500,000 EGP", "type": "بيع", "commission": 275000},
    "B15":      {"name": "Mr. Ehab B15 Properties", "price": "متعدد", "type": "بيع/إيجار", "commission": 200000},
    "PRIVADO":  {"name": "Privado (Studio + 2BR + Lake)", "price": "35K/mo + 5.5M + 2.5M", "type": "إيجار/بيع", "commission": 150000},
    "DREAMLAND":{"name": "Emeralds 3BR Dreamland", "price": "3M–5M EGP", "type": "إيجار تجاري", "commission": 250000},
}

KEYWORDS = {
    "B11":      ["b11","B11"],
    "B15":      ["b15","B15"],
    "PRIVADO":  ["privado","بريفادو","Privado"],
    "DREAMLAND":["dreamland","الأمارلدز","emerald","dream land"],
}

def wa_send(chat_id, message):
    cmd = ["curl","-s","-X","POST",
           f"https://7105.api.greenapi.com/waInstance{INSTANCE}/sendMessage/{TOKEN}",
           "-H","Content-Type: application/json",
           "-d", json.dumps({"chatId": chat_id, "message": message})]
    result = subprocess.run(cmd, capture_output=True, text=True)
    return "idMessage" in result.stdout

def wa_send_file(chat_id, file_url, filename, caption):
    cmd = ["curl","-s","-X","POST",
           f"https://7105.api.greenapi.com/waInstance{INSTANCE}/sendFileByUrl/{TOKEN}",
           "-H","Content-Type: application/json",
           "-d", json.dumps({"chatId": chat_id, "urlFile": file_url, "fileName": filename, "caption": caption})]
    result = subprocess.run(cmd, capture_output=True, text=True)
    return "idMessage" in result.stdout

def make_header(ws, title, subtitle, cols, color="1a5f7a"):
    ws.merge_cells(f"A1:{chr(64+cols)}1")
    t = ws["A1"]
    t.value = title
    t.font = Font(bold=True, size=13, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=color)
    t.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    ws.merge_cells(f"A2:{chr(64+cols)}2")
    s = ws["A2"]
    s.value = subtitle
    s.font = Font(bold=True, size=10, color="FFFFFF")
    s.fill = PatternFill("solid", fgColor="2980b9")
    s.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 25

def make_row_headers(ws, headers, row=3, color="1a5f7a"):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=color)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 28

def match_and_build(excel_path):
    wb = openpyxl.load_workbook(excel_path)
    ws_src = wb.active
    today = datetime.now().strftime("%Y-%m-%d %H:%M")
    date_str = datetime.now().strftime("%Y%m%d_%H%M")

    matched = {k: [] for k in CPI_ASSETS}
    seen = {k: set() for k in CPI_ASSETS}
    total = 0

    for row in ws_src.iter_rows(min_row=2, values_only=True):
        if not any(row): continue
        row_text = " ".join([str(c) for c in row if c])
        total += 1
        for asset_key, kws in KEYWORDS.items():
            for kw in kws:
                if kw in row_text:
                    phone = str(row[2]) if len(row)>2 and row[2] else ""
                    if phone and phone not in seen[asset_key]:
                        seen[asset_key].add(phone)
                        matched[asset_key].append(row)
                    break

    # Build output workbook
    out = openpyxl.Workbook()
    out.remove(out.active)

    HEADERS = ["#", "الاسم", "الهاتف", "واتساب", "نوع", "الميزانية", "غرف", "التاريخ", "الطلب", "الأولوية"]
    COLS = "ABCDEFGHIJ"

    summary_rows = []

    for asset_key, rows in matched.items():
        if not rows: continue
        asset = CPI_ASSETS[asset_key]
        ws = out.create_sheet(f"{asset_key} ({len(rows)})")

        make_header(ws, f"🏠 {asset['name']} — طلبات اليوم {today}",
                    f"إجمالي المطابقات: {len(rows)} | النوع: {asset['type']} | السعر: {asset['price']} | العمولة المتوقعة: {asset['commission']:,} EGP",
                    len(HEADERS))
        make_row_headers(ws, HEADERS)

        widths = [5,22,16,28,14,14,7,12,50,14]
        for i, w in enumerate(widths): ws.column_dimensions[COLS[i]].width = w

        for i, row in enumerate(rows):
            bg = "FFF3CD" if i < 3 else "D4EDDA" if i < 8 else "F0F8FF"
            priority = "🔥 الآن" if i < 3 else "⚡ اليوم" if i < 8 else "📋 الأسبوع"
            data = [i+1, row[1], row[2], row[3], row[5] if len(row)>5 else "", row[6] if len(row)>6 else "",
                    row[7] if len(row)>7 else "", row[9] if len(row)>9 else "",
                    str(row[8])[:120] if len(row)>8 else "", priority]
            for col, val in enumerate(data, 1):
                c = ws.cell(row=i+4, column=col, value=val)
                c.fill = PatternFill("solid", fgColor=bg.replace("#",""))
                c.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
                c.font = Font(size=9)
            ws.row_dimensions[i+4].height = 55

        summary_rows.append((asset_key, asset['name'], len(rows), asset['type'], asset['price'], asset['commission']))

    # Summary sheet
    ws_sum = out.create_sheet("📊 ملخص المطابقة", 0)
    make_header(ws_sum, f"📊 ملخص مطابقة عقارات CPI — {today}",
                f"إجمالي المشترين في السوق: {total} | إجمالي المطابقات: {sum(len(v) for v in matched.values())}",
                6, "7B2C2C")
    make_row_headers(ws_sum, ["العقار", "الاسم", "المطابقات", "النوع", "السعر", "العمولة المتوقعة"], color="7B2C2C")
    for col_letter, width in zip("ABCDEF", [12,30,14,14,22,22]):
        ws_sum.column_dimensions[col_letter].width = width

    total_commission = 0
    for i, (key, name, cnt, typ, price, comm) in enumerate(summary_rows):
        bg = "FFF3CD" if cnt >= 10 else "D4EDDA" if cnt >= 5 else "F0F8FF"
        row_data = [key, name, cnt, typ, price, f"{comm:,} EGP"]
        for col, val in enumerate(row_data, 1):
            c = ws_sum.cell(row=i+4, column=col, value=val)
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.font = Font(size=10, bold=(col==3))
        ws_sum.row_dimensions[i+4].height = 30
        total_commission += comm * min(cnt, 3)  # conservative: close 3 per asset max

    # Total row
    last = len(summary_rows) + 4
    ws_sum.merge_cells(f"A{last}:E{last}")
    tc = ws_sum.cell(row=last, column=1, value=f"💰 إجمالي العمولة المتوقعة (محافظ): {total_commission:,} EGP")
    tc.font = Font(bold=True, size=12, color="FFFFFF")
    tc.fill = PatternFill("solid", fgColor="1a5f7a")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[last].height = 35

    # Save
    filename = f"CPI_PropertyMatch_{date_str}.xlsx"
    out_path = f"{WEBROOT}/{filename}"
    out.save(out_path)
    print(f"✅ Saved: {out_path}")

    # Build WhatsApp summary
    match_lines = "\n".join([f"├─ {k}: {len(v)} مشتري" for k,v in matched.items() if v])
    total_matches = sum(len(v) for v in matched.values())

    message = (
        f"🏠 *تقرير مطابقة العقارات التلقائي*\n"
        f"📅 {today}\n\n"
        f"📊 *إجمالي السوق:* {total} مشتري نشط\n"
        f"✅ *مطابقات CPI:* {total_matches}\n\n"
        f"{match_lines}\n\n"
        f"💰 *العمولة المتوقعة:* {total_commission:,} EGP\n\n"
        f"📁 *الملف مرفق — 3 أولويات لكل عقار محددة*"
    )

    file_url = f"{BASE_URL}/{filename}"
    wa_send_file(MOMEN, file_url, filename, message)
    print(f"✅ Sent to Mo'men: {message[:80]}...")
    return total_matches, total_commission

if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else None
    if not path:
        # Find latest MatchPro demand file
        import glob
        files = sorted(glob.glob("/home/work/.openclaw/media/inbound/MatchPro_Demand*.xlsx"))
        path = files[-1] if files else None
    if path:
        print(f"Processing: {path}")
        matches, commission = match_and_build(path)
        print(f"✅ Done: {matches} matches | {commission:,} EGP commission")
    else:
        print("❌ No Excel file found")
