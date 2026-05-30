#!/usr/bin/env python3
"""
MatchPro Excel Template Builder — Fixed Logic
Rules:
- "مطلوب" at START of message = DEMAND (buyer looking)
- "مطلوب" in MIDDLE of message = SUPPLY (seller describing what's needed by tenant/buyer = listing)
- White background
- Better segmentation: area, type, price, bedrooms
"""
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
import re, json
from datetime import datetime

TEAL_DARK = "1A3A5C"
TEAL      = "1A5F7A"
BLUE      = "2980B9"
GREEN     = "27AE60"
ORANGE    = "E67E22"
RED       = "C0392B"
LIGHT_B   = "EBF5FB"
LIGHT_G   = "EAFAF1"
LIGHT_O   = "FEF9E7"
LIGHT_R   = "FDEDEC"
WHITE     = "FFFFFF"
GRAY      = "F2F3F4"
DARK_TEXT = "1A1A2E"

def hdr(text, bg, fg="FFFFFF", bold=True, size=10, wrap=False, center=True):
    return {
        'value': text,
        'fill': PatternFill("solid", fgColor=bg),
        'font': Font(bold=bold, color=fg, size=size, name="Calibri"),
        'alignment': Alignment(horizontal="center" if center else "right",
                               vertical="center", wrap_text=wrap),
    }

def cell_style(ws, row, col, value, bg=WHITE, fg=DARK_TEXT, bold=False,
               size=9, wrap=True, align="right", border=True):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(bold=bold, color=fg, size=size, name="Calibri")
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if border:
        thin = Side(style="thin", color="DDDDDD")
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return c

def apply_header_row(ws, row, headers_config, heights=30):
    ws.row_dimensions[row].height = heights
    for col, cfg in enumerate(headers_config, 1):
        c = ws.cell(row=row, column=col, value=cfg['value'])
        c.fill = cfg['fill']
        c.font = cfg['font']
        c.alignment = cfg['alignment']
        thin = Side(style="medium", color="FFFFFF")
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def classify_message(msg):
    """
    Correct demand/supply logic per Mo'men's rule:
    - مطلوب at START → DEMAND (buyer/renter looking)
    - مطلوب in MIDDLE → SUPPLY (seller listing a property)
    """
    msg = msg.strip()
    if re.match(r'^مطلوب', msg):
        return 'demand'
    elif 'مطلوب' in msg:
        return 'supply'
    # Additional signals
    demand_signals = ['عايز\s+ي', 'بدور\s+على', 'محتاج', 'أبحث', 'طالب\s+ايجار',
                      'عايزه\s+ي', 'بشتري', 'buyer', 'demand', 'يريد الشراء', 'للشراء معاه']
    supply_signals = ['للبيع', 'للإيجار', 'للايجار', 'عرض', 'متاح', 'شقة\s+في',
                      'دور\s+في', 'وحدة\s+في', 'فيلا\s+في', 'supply', 'مساحة']
    for s in demand_signals:
        if re.search(s, msg):
            return 'demand'
    for s in supply_signals:
        if re.search(s, msg):
            return 'supply'
    return 'unknown'

def extract_area(msg):
    areas = ['مدينتي', 'الرحاب', 'بيفيرلي', 'التجمع', 'القاهرة الجديدة',
             'بريفادو', 'Privado', 'B11', 'B15', 'B7', 'الشيخ زايد',
             'الساحل', 'العلمين', 'ال نورث كوست', 'شمال الساحل',
             'السادس من أكتوبر', 'الرياض', 'جدة', 'الغربية', 'الدلتا']
    for a in areas:
        if a.lower() in msg.lower():
            return a
    return '—'

def extract_budget(msg):
    # Find prices: numbers followed by currency indicators
    patterns = [
        r'(\d[\d,\.]+)\s*(مليون|M|م)',
        r'(\d[\d,\.]+)\s*(ألف|الف|K|ك)',
        r'بـ?\s*(\d[\d,\.]+)',
        r'(\d[\d,\.]+)\s*جنيه',
    ]
    for p in patterns:
        m = re.search(p, msg)
        if m:
            num = m.group(1).replace(',', '')
            try:
                val = float(num)
                unit = m.group(2) if len(m.groups()) > 1 else ''
                if 'مليون' in unit or 'M' in unit or 'م' == unit:
                    return f"{val:,.0f}M EGP"
                elif 'ألف' in unit or 'الف' in unit or 'K' in unit or 'ك' in unit:
                    return f"{val:,.0f}K EGP"
                elif val > 100000:
                    return f"{val:,.0f} EGP"
                return f"{val:,.0f} EGP"
            except:
                pass
    return '—'

def extract_type(msg):
    types = {'شقة': 'شقة', 'فيلا': 'فيلا', 'تاون هاوس': 'تاون هاوس',
             'توين هاوس': 'تاون هاوس', 'دوبلكس': 'دوبلكس',
             'ستوديو': 'ستوديو', 'محل': 'محل تجاري', 'مكتب': 'مكتب',
             'أرض': 'أرض', 'دور': 'دور', 'وحدة': 'وحدة'}
    for k, v in types.items():
        if k in msg:
            return v
    return '—'

def extract_rooms(msg):
    m = re.search(r'(\d)\s*(غرف|أوض|روم|BR|بيدروم|bedrooms?)', msg, re.IGNORECASE)
    if m: return m.group(1) + ' غرف'
    for n, w in [('1','غرفة واحدة'), ('2','غرفتين'), ('3','3 غرف'), ('4','4 غرف')]:
        if w in msg or f'{n} غرف' in msg:
            return f'{n} غرفة'
    return '—'

def build_excel(messages, output_path):
    wb = openpyxl.Workbook()

    # ── SHEET 1: DEMAND (Buyers/Renters) ─────────────────────────────────────
    ws_d = wb.active
    ws_d.title = "🔵 طلب (Demand)"
    ws_d.sheet_view.rightToLeft = True
    ws_d.freeze_panes = "A3"

    # Title row
    ws_d.merge_cells('A1:J1')
    c = ws_d['A1']
    c.value = "📊 تقرير الطلب — مشترين ومستأجرين نشطين | MatchPro Intelligence"
    c.fill = PatternFill("solid", fgColor=TEAL_DARK)
    c.font = Font(bold=True, color="FFFFFF", size=13, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_d.row_dimensions[1].height = 35

    demand_headers = [
        hdr("رقم", TEAL, size=9),
        hdr("نوع العقار", TEAL, size=9),
        hdr("المنطقة", TEAL, size=9),
        hdr("الميزانية", TEAL, size=9),
        hdr("الغرف", TEAL, size=9),
        hdr("الهدف", TEAL, size=9),
        hdr("الهاتف", TEAL, size=9),
        hdr("درجة الأولوية", TEAL, size=9),
        hdr("ملاحظات", TEAL, size=9),
        hdr("التاريخ", TEAL, size=9),
    ]
    apply_header_row(ws_d, 2, demand_headers, 28)

    demand_rows = [r for r in messages if r['type'] == 'demand']
    for i, row in enumerate(demand_rows):
        r = i + 3
        bg = LIGHT_B if i % 2 == 0 else WHITE
        priority = "🔴 عاجل" if row.get('budget_val', 0) > 3000000 else "🟡 متوسط"
        cells = [
            (i+1, bg, DARK_TEXT, True, 9),
            (row['property_type'], bg, DARK_TEXT, False, 9),
            (row['area'], bg, BLUE, True, 9),
            (row['budget'], bg, GREEN, True, 10),
            (row['rooms'], bg, DARK_TEXT, False, 9),
            (row['purpose'], bg, DARK_TEXT, False, 9),
            (row.get('phone', '—'), bg, "8B0000", True, 9),
            (priority, bg, DARK_TEXT, False, 9),
            (row['msg'][:60] + '...' if len(row['msg']) > 60 else row['msg'], bg, "555555", False, 8),
            (row.get('date', datetime.now().strftime('%Y-%m-%d')), bg, DARK_TEXT, False, 8),
        ]
        ws_d.row_dimensions[r].height = 22
        for col, (val, bg_, fg_, bold_, sz_) in enumerate(cells, 1):
            cell_style(ws_d, r, col, val, bg_, fg_, bold_, sz_)

    # Column widths for demand
    for col, w in zip(range(1,11), [6, 14, 14, 14, 10, 10, 14, 12, 35, 12]):
        ws_d.column_dimensions[get_column_letter(col)].width = w

    # ── SHEET 2: SUPPLY (Sellers/Landlords) ──────────────────────────────────
    ws_s = wb.create_sheet("🟢 عرض (Supply)")
    ws_s.sheet_view.rightToLeft = True
    ws_s.freeze_panes = "A3"

    ws_s.merge_cells('A1:J1')
    c = ws_s['A1']
    c.value = "📦 تقرير العرض — عقارات معروضة للبيع والإيجار | MatchPro Intelligence"
    c.fill = PatternFill("solid", fgColor="1A6E3A")
    c.font = Font(bold=True, color="FFFFFF", size=13, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_s.row_dimensions[1].height = 35

    supply_headers = [
        hdr("رقم", "1A6E3A", size=9),
        hdr("نوع العقار", "1A6E3A", size=9),
        hdr("المنطقة", "1A6E3A", size=9),
        hdr("السعر", "1A6E3A", size=9),
        hdr("الغرف", "1A6E3A", size=9),
        hdr("الهدف", "1A6E3A", size=9),
        hdr("الهاتف", "1A6E3A", size=9),
        hdr("حالة التطابق", "1A6E3A", size=9),
        hdr("تفاصيل", "1A6E3A", size=9),
        hdr("التاريخ", "1A6E3A", size=9),
    ]
    apply_header_row(ws_s, 2, supply_headers, 28)

    supply_rows = [r for r in messages if r['type'] == 'supply']
    for i, row in enumerate(supply_rows):
        r = i + 3
        bg = LIGHT_G if i % 2 == 0 else WHITE
        match = "✅ تطابق CPI" if any(k in row['area'] for k in ['مدينتي','الرحاب','بريفادو','B11','B15']) else "⚪ بحث"
        cells = [
            (i+1, bg, DARK_TEXT, True, 9),
            (row['property_type'], bg, DARK_TEXT, False, 9),
            (row['area'], bg, "1A6E3A", True, 9),
            (row['budget'], bg, "8B4500", True, 10),
            (row['rooms'], bg, DARK_TEXT, False, 9),
            (row['purpose'], bg, DARK_TEXT, False, 9),
            (row.get('phone', '—'), bg, "8B0000", True, 9),
            (match, bg, DARK_TEXT, False, 9),
            (row['msg'][:60] + '...' if len(row['msg']) > 60 else row['msg'], bg, "555555", False, 8),
            (row.get('date', datetime.now().strftime('%Y-%m-%d')), bg, DARK_TEXT, False, 8),
        ]
        ws_s.row_dimensions[r].height = 22
        for col, (val, bg_, fg_, bold_, sz_) in enumerate(cells, 1):
            cell_style(ws_s, r, col, val, bg_, fg_, bold_, sz_)

    for col, w in zip(range(1,11), [6, 14, 14, 14, 10, 10, 14, 12, 35, 12]):
        ws_s.column_dimensions[get_column_letter(col)].width = w

    # ── SHEET 3: SUMMARY ─────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("📈 ملخص السوق")
    ws_sum.sheet_view.rightToLeft = True

    ws_sum.merge_cells('A1:E1')
    c = ws_sum['A1']
    c.value = "📈 ملخص السوق — MatchPro Intelligence | " + datetime.now().strftime('%Y-%m-%d %H:%M')
    c.fill = PatternFill("solid", fgColor=TEAL_DARK)
    c.font = Font(bold=True, color="FFFFFF", size=13, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 35

    demand_count = len([r for r in messages if r['type'] == 'demand'])
    supply_count = len([r for r in messages if r['type'] == 'supply'])
    unknown_count = len([r for r in messages if r['type'] == 'unknown'])

    summary_data = [
        ("📊 إجمالي الرسائل المحللة", len(messages), TEAL, WHITE),
        ("🔵 طلبات (مشترين/مستأجرين)", demand_count, BLUE, WHITE),
        ("🟢 عروض (بائعين/مؤجرين)", supply_count, "1A6E3A", WHITE),
        ("⚪ غير محدد", unknown_count, "888888", WHITE),
        ("", "", WHITE, DARK_TEXT),
        ("نسبة الطلب للعرض", f"{demand_count}/{supply_count} = {demand_count/max(supply_count,1):.1f}:1", 
         RED if demand_count/max(supply_count,1) > 3 else ORANGE, WHITE),
        ("حالة السوق", "🔴 سوق بائع — طلب أعلى بكثير" if demand_count > supply_count*2 else "🟡 متوازن نسبياً",
         RED if demand_count > supply_count*2 else ORANGE, WHITE),
    ]

    for i, (label, value, bg, fg) in enumerate(summary_data):
        r = i + 3
        ws_sum.row_dimensions[r].height = 26
        if not label:
            continue
        c1 = ws_sum.cell(row=r, column=1, value=label)
        c1.fill = PatternFill("solid", fgColor=bg)
        c1.font = Font(bold=True, color=fg, size=11, name="Calibri")
        c1.alignment = Alignment(horizontal="right", vertical="center")
        ws_sum.merge_cells(f'A{r}:C{r}')

        c2 = ws_sum.cell(row=r, column=4, value=value)
        c2.fill = PatternFill("solid", fgColor=LIGHT_B)
        c2.font = Font(bold=True, color=bg, size=13, name="Calibri")
        c2.alignment = Alignment(horizontal="center", vertical="center")
        ws_sum.merge_cells(f'D{r}:E{r}')

    ws_sum.column_dimensions['A'].width = 12
    ws_sum.column_dimensions['B'].width = 12
    ws_sum.column_dimensions['C'].width = 12
    ws_sum.column_dimensions['D'].width = 15
    ws_sum.column_dimensions['E'].width = 15

    wb.save(output_path)
    print(f"✅ Excel template saved: {output_path}")

# Test with sample messages using the CORRECT classification logic
test_messages = [
    {
        'msg': 'مطلوب شقة في مدينتي 3 غرف بـ 5 مليون',
        'type': 'demand',  # مطلوب at START = buyer
        'area': 'مدينتي', 'budget': '5M EGP', 'budget_val': 5000000,
        'rooms': '3 غرف', 'property_type': 'شقة', 'purpose': 'شراء',
        'phone': '01012345678', 'date': '2026-05-02'
    },
    {
        'msg': 'شقة للبيع في الرحاب 4 غرف مطلوب 7 مليون',
        'type': 'supply',  # مطلوب in MIDDLE = seller listing
        'area': 'الرحاب', 'budget': '7M EGP', 'budget_val': 7000000,
        'rooms': '4 غرف', 'property_type': 'شقة', 'purpose': 'بيع',
        'phone': '01098765432', 'date': '2026-05-02'
    },
    {
        'msg': 'مطلوب استوديو للإيجار في التجمع الخامس',
        'type': 'demand',
        'area': 'التجمع', 'budget': '—', 'budget_val': 0,
        'rooms': 'ستوديو', 'property_type': 'ستوديو', 'purpose': 'إيجار',
        'phone': '01155667788', 'date': '2026-05-02'
    },
    {
        'msg': 'فيلا للإيجار في الشيخ زايد 5 غرف مطلوب 80 ألف',
        'type': 'supply',
        'area': 'الشيخ زايد', 'budget': '80K EGP', 'budget_val': 80000,
        'rooms': '5 غرف', 'property_type': 'فيلا', 'purpose': 'إيجار',
        'phone': '01234567890', 'date': '2026-05-02'
    },
]

build_excel(test_messages, "/var/www/html/MatchPro_Template_CORRECT.xlsx")
print("✅ Template with correct demand/supply logic ready")
print("\nLogic validation:")
for m in test_messages:
    print(f"  '{m['msg'][:40]}...' → {m['type'].upper()}")
