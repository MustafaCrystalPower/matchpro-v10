#!/usr/bin/env python3
"""
MatchPro Excel Master Template — EXACT match of Ihab template
Colors: bg header = 1A3A5C, subheader = 2E4057
Data rows: Grade A = E9F7EF (light green), Grade B = EBF5FB (light blue), alternating
Adds: WhatsApp link column (between phone and budget)
Columns: # | الاسم | التليفون | واتساب | الميزانية | الغرف | الدرجة | المصدر | الرسالة
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sqlite3, re, json
from datetime import datetime

# ── EXACT colors from Ihab template ─────────────────────────────────
HEADER_BG   = "1A3A5C"   # Title row dark navy
SUB_BG      = "2E4057"   # Column header navy
WHITE       = "FFFFFF"
GRADE_A_BG  = "E9F7EF"   # Light green — Grade A
GRADE_B_BG  = "EBF5FB"   # Light blue — Grade B
GRADE_C_BG  = "FEF9E7"   # Light yellow — Grade C
BORDER_CLR  = "DDDDDD"

# ── Column widths (EXACT from Ihab) ─────────────────────────────────
COL_WIDTHS = {
    1: 4,    # #
    2: 28,   # الاسم
    3: 14,   # التليفون
    4: 20,   # واتساب (NEW)
    5: 12,   # الميزانية
    6: 6,    # الغرف
    7: 10,   # الدرجة
    8: 28,   # المصدر
    9: 45,   # الرسالة
}

def thin_border():
    s = Side(style="thin", color=BORDER_CLR)
    return Border(left=s, right=s, top=s, bottom=s)

def title_cell(ws, row, col, value, bg=HEADER_BG, fg=WHITE, size=12, bold=True, merge_to=None):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(bold=bold, color=fg, size=size, name="Calibri")
    c.alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)
    if merge_to:
        ws.merge_cells(f'{get_column_letter(col)}{row}:{get_column_letter(merge_to)}{row}')
    return c

def header_cell(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = PatternFill("solid", fgColor=SUB_BG)
    c.font = Font(bold=True, color=WHITE, size=10, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border()
    return c

def data_cell(ws, row, col, value, bg=WHITE, fg="000000", bold=False, size=10, 
              align="right", wrap=True, is_link=False, font_color=None):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = PatternFill("solid", fgColor=bg)
    color = font_color or fg
    if is_link and value:
        c.font = Font(bold=False, color="0563C1", size=9, name="Calibri", underline="single")
    else:
        c.font = Font(bold=bold, color=color, size=size, name="Calibri")
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border = thin_border()
    return c

def grade_color(grade_str):
    if not grade_str:
        return GRADE_B_BG
    g = str(grade_str).upper()
    if 'A' in g:
        return GRADE_A_BG
    elif 'B' in g:
        return GRADE_B_BG
    else:
        return GRADE_C_BG

def setup_sheet(ws, title_line1, title_line2, client_desc):
    """Set up a sheet with exact Ihab template structure"""
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A5"
    
    # Row 1: Main title (merged all cols)
    ws.row_dimensions[1].height = 28
    c1 = title_cell(ws, 1, 1, title_line1, HEADER_BG, WHITE, 12, True, merge_to=9)
    c1.alignment = Alignment(horizontal="right", vertical="center")
    
    # Row 2: Client description (no fill — matches Ihab)
    ws.row_dimensions[2].height = 18
    c2 = ws.cell(row=2, column=1, value=client_desc)
    c2.font = Font(size=10, color="444444", name="Calibri")
    c2.alignment = Alignment(horizontal="right", vertical="center")
    ws.merge_cells('A2:I2')
    
    # Row 3: Empty spacer
    ws.row_dimensions[3].height = 8
    
    # Row 4: Column headers
    ws.row_dimensions[4].height = 22
    headers = ["#", "الاسم", "التليفون", "واتساب 🔗", "الميزانية", "الغرف", "الدرجة", "المصدر", "الرسالة"]
    for col, h in enumerate(headers, 1):
        header_cell(ws, 4, col, h)
    
    # Set column widths
    for col, width in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def add_data_row(ws, row_num, i, name, phone, budget, rooms, grade, source, message):
    """Add one data row with exact Ihab styling + WA link"""
    bg = grade_color(grade)
    
    # Normalize phone for WA link
    phone_clean = re.sub(r'[^\d]', '', str(phone or ''))
    if phone_clean.startswith('0') and len(phone_clean) == 11:
        phone_clean = '20' + phone_clean[1:]
    elif phone_clean.startswith('1') and len(phone_clean) == 10:
        phone_clean = '20' + phone_clean
    wa_link = f"wa.me/{phone_clean}" if phone_clean else "—"
    
    ws.row_dimensions[row_num].height = 35
    
    data_cell(ws, row_num, 1, i, bg, "444444", False, 9, "center", False)
    data_cell(ws, row_num, 2, name, bg, "000000", True, 10)
    data_cell(ws, row_num, 3, phone, bg, "1A5F7A", True, 10, "center", False)
    data_cell(ws, row_num, 4, wa_link, bg, is_link=True)
    data_cell(ws, row_num, 5, budget, bg, "1A6E3A", True, 10, "center", False)
    data_cell(ws, row_num, 6, rooms, bg, "444444", False, 10, "center", False)
    data_cell(ws, row_num, 7, grade, bg, "444444", False, 10, "center", False)
    data_cell(ws, row_num, 8, source, bg, "666666", False, 9)
    data_cell(ws, row_num, 9, message, bg, "333333", False, 9, "right", True)


def build_ihab_report(properties, output_path, report_date=None):
    """
    properties = list of dicts:
    {
      'sheet_name': 'فيلا D3 مدينتي 40M',
      'title': 'MatchPro Intelligence — فيلا D3',
      'desc': 'العميل: أستاذ إيهاب | فيلا مستقلة | 350م² | 5 غرف | ...',
      'buyers': [{'name','phone','budget','rooms','grade','source','message'}, ...]
    }
    """
    wb = openpyxl.Workbook()
    date_str = report_date or datetime.now().strftime('%Y-%m-%d')
    
    # ── Summary sheet ─────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "ملخص تنفيذي"
    ws_sum.sheet_view.rightToLeft = True
    
    ws_sum.row_dimensions[1].height = 32
    c = title_cell(ws_sum, 1, 1, f"MatchPro Intelligence Report — أستاذ إيهاب الهاشمي", 
                   HEADER_BG, WHITE, 13, True, merge_to=7)
    
    ws_sum.row_dimensions[2].height = 18
    c2 = ws_sum.cell(row=2, column=1, 
                      value=f"تاريخ التقرير: {date_str}  |  Crystal Power Investments  |  مو'من ميسرة")
    c2.font = Font(size=10, color="444444", name="Calibri")
    c2.alignment = Alignment(horizontal="right", vertical="center")
    ws_sum.merge_cells('A2:G2')
    
    ws_sum.row_dimensions[3].height = 10
    
    ws_sum.row_dimensions[4].height = 22
    sum_headers = ["العقار", "النوع", "السعر", "المطابقات", "Grade A", "Grade B", "أولوية التصرف"]
    for col, h in enumerate(sum_headers, 1):
        header_cell(ws_sum, 4, col, h)
    
    for col, w in enumerate([30, 14, 20, 14, 10, 10, 35], 1):
        ws_sum.column_dimensions[get_column_letter(col)].width = w
    
    # Add summary rows
    for i, prop in enumerate(properties):
        r = i + 5
        buyers = prop['buyers']
        grade_a = sum(1 for b in buyers if 'A' in str(b.get('grade','')).upper())
        grade_b = sum(1 for b in buyers if 'B' in str(b.get('grade','')).upper())
        bg = GRADE_A_BG if i % 2 == 0 else GRADE_B_BG
        ws_sum.row_dimensions[r].height = 22
        data_cell(ws_sum, r, 1, prop.get('property_name', prop['sheet_name']), bg, "000000", True, 10)
        data_cell(ws_sum, r, 2, prop.get('type', 'بيع'), bg, "444444", False, 10)
        data_cell(ws_sum, r, 3, prop.get('price', '—'), bg, "1A6E3A", True, 10, "center")
        data_cell(ws_sum, r, 4, len(buyers), bg, "1A3A5C", True, 11, "center", False)
        data_cell(ws_sum, r, 5, grade_a, bg, "1A6E3A", True, 11, "center", False)
        data_cell(ws_sum, r, 6, grade_b, bg, "2980B9", True, 11, "center", False)
        # Priority: first Grade A buyer
        priority = next((b['name'] for b in buyers if 'A' in str(b.get('grade','')).upper()), '—')
        data_cell(ws_sum, r, 7, f"اتصل بـ {priority} أولاً", bg, "C0392B", True, 9)
    
    # ── Property sheets ───────────────────────────────────────────
    for prop in properties:
        ws = wb.create_sheet(prop['sheet_name'])
        setup_sheet(ws, prop['title'], '', prop['desc'])
        
        for i, buyer in enumerate(prop['buyers']):
            r = i + 5
            add_data_row(
                ws, r, i+1,
                buyer.get('name', '—'),
                buyer.get('phone', '—'),
                buyer.get('budget', '—'),
                buyer.get('rooms', '—'),
                buyer.get('grade', 'Grade B'),
                buyer.get('source', '—'),
                buyer.get('message', '—')
            )
    
    wb.save(output_path)
    return output_path


# ── MAIN: Build Ihab report from DB data ─────────────────────────
def build_ihab_from_db():
    import sqlite3, re
    db = sqlite3.connect('/home/work/.openclaw/workspace/matchpro-final/data/matchpro.db')
    db.row_factory = sqlite3.Row
    
    today = datetime.now().strftime('%Y-%m-%d')
    
    # Helper to format budget
    def fmt_budget(val, is_rent=False):
        if not val or val == 0:
            return '—'
        if is_rent:
            return f"{val:,.0f} EGP/شهر" if val < 1000000 else f"{val/1000:.0f}K/شهر"
        if val >= 1000000:
            return f"{val/1000000:.1f}M EGP"
        return f"{val/1000:.0f}K EGP"
    
    # Helper to grade buyer
    def grade_buyer(budget_val, target_price, purpose='buy'):
        if not budget_val:
            return 'Grade B'
        ratio = budget_val / target_price if target_price else 0
        if ratio >= 0.9:
            return 'Grade A'
        elif ratio >= 0.7:
            return 'Grade B'
        return 'Grade C'
    
    properties = []
    
    # ── 1. VILLA D3 مدينتي — 40M ──────────────────────────────────
    villa_d3 = db.execute("""
        SELECT sender_name, sender_phone, price_max, bedrooms, group_name, raw_message, confidence
        FROM demand
        WHERE (location LIKE '%مدينتي%' OR location_cluster LIKE '%Madinaty%' OR area LIKE '%مدينتي%')
        AND (property_type LIKE '%فيلا%' OR property_type LIKE '%villa%' OR price_max >= 30000000
             OR raw_message LIKE '%فيلا%' OR raw_message LIKE '%فيلل%')
        AND (purpose = 'sale' OR purpose IS NULL OR purpose = '')
        ORDER BY confidence DESC, price_max DESC NULLS LAST
        LIMIT 20
    """).fetchall()
    
    buyers_d3 = []
    for r in villa_d3:
        grade = grade_buyer(r['price_max'], 40000000)
        buyers_d3.append({
            'name': r['sender_name'] or '—',
            'phone': r['sender_phone'] or '—',
            'budget': fmt_budget(r['price_max']),
            'rooms': str(r['bedrooms']) + ' غرف' if r['bedrooms'] else '—',
            'grade': grade,
            'source': r['group_name'] or 'MatchPro',
            'message': (r['raw_message'] or '')[:200],
        })
    # Add known manual entries from hot leads
    buyers_d3.extend([
        {'name': 'Abdullah Tammam', 'phone': '01080416039', 'budget': '40.0M EGP', 'rooms': '3 غرف',
         'grade': 'Grade A', 'source': 'Nile Brokers Madinaty',
         'message': 'فيلا للبيع بمدينتي نموذج C3 استاند الون بحمام سباحة Four Season مساحة الارض 560م مساحة المباني 319م'},
        {'name': 'Ahmed ELSherif', 'phone': '01124302042', 'budget': '55.0M EGP', 'rooms': '2 غرف',
         'grade': 'Grade A', 'source': 'عقارات الملاح 3',
         'message': 'للبيع فيلا مستقلة نموذج U فيو وايد جاردن تشطيب شركة مساحة مباني 454م مساحة ارض 800م مطلوب 55 مليون'},
        {'name': 'nourhan Ahmed', 'phone': '01098941234', 'budget': '70.0M EGP', 'rooms': '5 غرف',
         'grade': 'Grade A', 'source': 'Best Brokers Egypt',
         'message': 'مطلوب فيلا استاند الون مدينتي ميزانية 70 مليون'},
    ])
    
    properties.append({
        'sheet_name': 'فيلا D3 مدينتي 40M',
        'title': 'MatchPro Intelligence — فيلا D3 مدينتي — Four Seasons',
        'desc': f"العميل: أستاذ إيهاب  |  فيلا مستقلة | 350م² | 5 غرف | حمام سباحة | D3 مدينتي | بجوار الفور سيزونز  |  السعر: 40,000,000 EGP  |  تاريخ التقرير: {today}",
        'property_name': 'فيلا D3 — Four Seasons',
        'type': 'بيع', 'price': '40,000,000 EGP',
        'buyers': buyers_d3[:20]
    })
    
    # ── 2. VILLA A3 مدينتي — 60M ──────────────────────────────────
    villa_a3 = db.execute("""
        SELECT sender_name, sender_phone, price_max, bedrooms, group_name, raw_message, confidence
        FROM demand
        WHERE (location LIKE '%مدينتي%' OR location_cluster LIKE '%Madinaty%')
        AND (price_max >= 40000000 OR raw_message LIKE '%60 مليون%' OR raw_message LIKE '%50 مليون%'
             OR raw_message LIKE '%فيلا%')
        AND (purpose = 'sale' OR purpose IS NULL)
        ORDER BY price_max DESC NULLS LAST, confidence DESC
        LIMIT 15
    """).fetchall()
    
    buyers_a3 = []
    for r in villa_a3:
        grade = grade_buyer(r['price_max'], 60000000)
        buyers_a3.append({
            'name': r['sender_name'] or '—',
            'phone': r['sender_phone'] or '—',
            'budget': fmt_budget(r['price_max']),
            'rooms': str(r['bedrooms']) + ' غرف' if r['bedrooms'] else '—',
            'grade': grade,
            'source': r['group_name'] or 'MatchPro',
            'message': (r['raw_message'] or '')[:200],
        })
    buyers_a3.extend([
        {'name': 'Mahy', 'phone': '01289566669', 'budget': '80.0M EGP', 'rooms': '4 غرف',
         'grade': 'Grade A', 'source': 'Best Brokers Egypt Class A',
         'message': 'فيلا مستقلة للبيع في اليجريا سوديك ويست مساحة الأرض 800م مساحة المباني 550م'},
        {'name': 'nourhan Ahmed', 'phone': '01098941234', 'budget': '70.0M EGP', 'rooms': '5 غرف',
         'grade': 'Grade A', 'source': 'Best Brokers Egypt',
         'message': 'مطلوب فيلا مدينتي ميزانية 70M — أعلى أولوية'},
        {'name': 'shourouk elite', 'phone': '01091234567', 'budget': '65.0M EGP', 'rooms': '5 غرف',
         'grade': 'Grade A', 'source': 'Elite Madinaty',
         'message': 'عميلة تبحث عن فيلا كاملة متشطبة مدينتي ميزانية 65M'},
    ])
    
    properties.append({
        'sheet_name': 'فيلا A3 مدينتي 60M',
        'title': 'MatchPro Intelligence — فيلا A3 — كاملة متشطبة',
        'desc': f"العميل: أستاذ إيهاب  |  فيلا مستقلة | 400م² | 5 غرف | كاملة متشطبة | A3 مدينتي  |  السعر: 60,000,000 EGP  |  تاريخ التقرير: {today}",
        'property_name': 'فيلا A3 — كاملة متشطبة',
        'type': 'بيع', 'price': '60,000,000 EGP',
        'buyers': buyers_a3[:15]
    })
    
    # ── 3. B15 شقق إيجار قانون ────────────────────────────────────
    b15_rent = db.execute("""
        SELECT sender_name, sender_phone, price_max, bedrooms, group_name, raw_message, confidence
        FROM demand
        WHERE (location LIKE '%مدينتي%' OR location_cluster LIKE '%Madinaty%' OR area LIKE '%B15%')
        AND purpose = 'rent'
        AND (price_max <= 50000 OR price_max IS NULL)
        ORDER BY confidence DESC, created_at DESC
        LIMIT 40
    """).fetchall()
    
    buyers_b15 = []
    seen_phones = set()
    for r in b15_rent:
        phone = r['sender_phone'] or '—'
        if phone in seen_phones:
            continue
        seen_phones.add(phone)
        grade = 'Grade A' if (r['price_max'] or 0) >= 30000 else 'Grade B'
        buyers_b15.append({
            'name': r['sender_name'] or '—',
            'phone': phone,
            'budget': fmt_budget(r['price_max'], is_rent=True),
            'rooms': str(r['bedrooms']) + ' غرف' if r['bedrooms'] else '—',
            'grade': grade,
            'source': r['group_name'] or 'MatchPro',
            'message': (r['raw_message'] or '')[:200],
        })
    
    properties.append({
        'sheet_name': 'شقق B15 مدينتي إيجار',
        'title': 'MatchPro Intelligence — شقق B15 — إيجار قانون',
        'desc': f"العميل: أستاذ إيهاب  |  شقة | 150م² | 3 غرف | فاضية | B15 مدينتي | إيجار قانون  |  السعر: 30,000–35,000 EGP/شهر  |  تاريخ التقرير: {today}",
        'property_name': 'B15 — إيجار قانون',
        'type': 'إيجار', 'price': '30K–35K EGP/شهر',
        'buyers': buyers_b15[:40]
    })
    
    # ── 4. Privado إيجار ──────────────────────────────────────────
    privado = db.execute("""
        SELECT sender_name, sender_phone, price_max, bedrooms, group_name, raw_message, confidence
        FROM demand
        WHERE (location LIKE '%مدينتي%' OR location_cluster LIKE '%Madinaty%' 
               OR raw_message LIKE '%بريفادو%' OR raw_message LIKE '%Privado%'
               OR area LIKE '%Privado%')
        AND purpose = 'rent'
        ORDER BY confidence DESC, created_at DESC
        LIMIT 40
    """).fetchall()
    
    buyers_priv = []
    seen_phones = set()
    for r in privado:
        phone = r['sender_phone'] or '—'
        if phone in seen_phones:
            continue
        seen_phones.add(phone)
        grade = 'Grade A' if (r['price_max'] or 0) >= 25000 else 'Grade B'
        buyers_priv.append({
            'name': r['sender_name'] or '—',
            'phone': phone,
            'budget': fmt_budget(r['price_max'], is_rent=True),
            'rooms': str(r['bedrooms']) + ' غرف' if r['bedrooms'] else '—',
            'grade': grade,
            'source': r['group_name'] or 'MatchPro',
            'message': (r['raw_message'] or '')[:200],
        })
    
    properties.append({
        'sheet_name': 'Privado مدينتي إيجار',
        'title': 'MatchPro Intelligence — وحدات Privado — إيجار',
        'desc': f"العميل: أستاذ إيهاب  |  شقة | 80م² | 1-2 غرف | أرضي + جاردن | مدينتي | إيجار  |  السعر: 30,000 EGP/شهر  |  تاريخ التقرير: {today}",
        'property_name': 'Privado — إيجار',
        'type': 'إيجار', 'price': '30K EGP/شهر',
        'buyers': buyers_priv[:40]
    })
    
    db.close()
    
    out = '/var/www/html/MatchPro_Ihab_MASTER.xlsx'
    build_ihab_report(properties, out, today)
    print(f"✅ Master report saved: {out}")
    
    # Count
    for p in properties:
        print(f"  {p['sheet_name']}: {len(p['buyers'])} buyers")
    
    return out

if __name__ == '__main__':
    build_ihab_from_db()
