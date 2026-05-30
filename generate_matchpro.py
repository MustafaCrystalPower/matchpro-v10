#!/usr/bin/env python3
"""MatchPro Investor Package Generator — English Only"""

import os
import sys

# ── dependencies ──────────────────────────────────────────────────────────────
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm, cm
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                     TableStyle, HRFlowable, PageBreak, KeepTogether)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
    from reportlab.pdfgen import canvas as pdfcanvas
    from reportlab.platypus.flowables import Flowable
except ImportError:
    print("Installing reportlab...")
    os.system("pip3 install reportlab --break-system-packages -q")
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm, cm
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                     TableStyle, HRFlowable, PageBreak, KeepTogether)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
    from reportlab.pdfgen import canvas as pdfcanvas
    from reportlab.platypus.flowables import Flowable

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    os.system("pip3 install openpyxl --break-system-packages -q")
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# ── colours ───────────────────────────────────────────────────────────────────
NAVY  = colors.HexColor("#0B2545")
GOLD  = colors.HexColor("#C9A961")
WHITE = colors.white
RED   = colors.HexColor("#CC2222")
LGRAY = colors.HexColor("#F5F5F5")
LBLUE = colors.HexColor("#F0F4FF")
GOLD_LIGHT = colors.HexColor("#FFF8E7")
NAVY_LIGHT = colors.HexColor("#E8EDF5")

PAGE_W, PAGE_H = A4
MARGIN = 18*mm

# ── helpers ───────────────────────────────────────────────────────────────────

def style(name="Normal", **kw):
    s = ParagraphStyle(name)
    for k, v in kw.items():
        setattr(s, k, v)
    return s

def P(text, **kw):
    return Paragraph(text, style(**kw))

def navy_header(text, sub=None):
    """Dark navy header strip with optional subtitle."""
    elems = []
    elems.append(Spacer(1, 4*mm))
    header_data = [[Paragraph(text, style(
        "h", fontSize=20, textColor=WHITE, fontName="Helvetica-Bold",
        leading=24, leftIndent=6, spaceAfter=0, spaceBefore=0))]]
    if sub:
        header_data.append([Paragraph(sub, style(
            "s", fontSize=10, textColor=GOLD, fontName="Helvetica",
            leading=14, leftIndent=6, spaceAfter=0, spaceBefore=0))])
    t = Table(header_data, colWidths=[PAGE_W - 2*MARGIN])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), NAVY),
        ("TOPPADDING", (0,0), (-1,-1), 8),
        ("BOTTOMPADDING", (0,0), (-1,-1), 8),
        ("LEFTPADDING", (0,0), (-1,-1), 10),
        ("RIGHTPADDING", (0,0), (-1,-1), 10),
    ]))
    elems.append(t)
    elems.append(Spacer(1, 5*mm))
    return elems

def red_box(title, body):
    """Problem box with red left border."""
    inner = [
        [Paragraph(f"<b>{title}</b>", style("rb_t", fontSize=11, textColor=NAVY,
                   fontName="Helvetica-Bold", leading=14, spaceAfter=2)),
         ""],
        [Paragraph(body, style("rb_b", fontSize=9.5, textColor=colors.HexColor("#222222"),
                   fontName="Helvetica", leading=13, spaceAfter=0)),
         ""],
    ]
    # Red left-border trick: thin red column
    t = Table([[
        Table([[""]], colWidths=[4], rowHeights=[None],
              style=TableStyle([("BACKGROUND",(0,0),(-1,-1), RED),
                                ("TOPPADDING",(0,0),(-1,-1),0),
                                ("BOTTOMPADDING",(0,0),(-1,-1),0),
                                ("LEFTPADDING",(0,0),(-1,-1),0),
                                ("RIGHTPADDING",(0,0),(-1,-1),0)])),
        Table(inner, colWidths=[PAGE_W - 2*MARGIN - 14, 0],
              style=TableStyle([
                  ("SPAN",(0,0),(1,0)),
                  ("SPAN",(0,1),(1,1)),
                  ("TOPPADDING",(0,0),(-1,-1),4),
                  ("BOTTOMPADDING",(0,0),(-1,-1),4),
                  ("LEFTPADDING",(0,0),(-1,-1),8),
                  ("RIGHTPADDING",(0,0),(-1,-1),4),
                  ("VALIGN",(0,0),(-1,-1),"TOP"),
              ]))
    ]], colWidths=[4, PAGE_W - 2*MARGIN - 4],
    style=TableStyle([
        ("BACKGROUND",(0,0),(-1,-1), colors.HexColor("#FFF5F5")),
        ("TOPPADDING",(0,0),(-1,-1),0),
        ("BOTTOMPADDING",(0,0),(-1,-1),0),
        ("LEFTPADDING",(0,0),(-1,-1),0),
        ("RIGHTPADDING",(0,0),(-1,-1),0),
        ("BOX",(0,0),(-1,-1),1,colors.HexColor("#FFCCCC")),
    ]))
    return [t, Spacer(1, 3*mm)]

def navy_feature_box(title, body):
    """Navy feature box white text."""
    data = [
        [Paragraph(f"<b>{title}</b>", style("nf_t", fontSize=11, textColor=GOLD,
                   fontName="Helvetica-Bold", leading=14))],
        [Paragraph(body, style("nf_b", fontSize=9.5, textColor=WHITE,
                   fontName="Helvetica", leading=13))],
    ]
    t = Table(data, colWidths=[PAGE_W - 2*MARGIN])
    t.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,-1), NAVY),
        ("TOPPADDING",(0,0),(-1,-1),8),
        ("BOTTOMPADDING",(0,0),(-1,-1),8),
        ("LEFTPADDING",(0,0),(-1,-1),10),
        ("RIGHTPADDING",(0,0),(-1,-1),10),
        ("ROWBACKGROUNDS",(0,0),(-1,-1),[NAVY]),
    ]))
    return [t, Spacer(1, 3*mm)]

def stat_grid(stats):
    """2-row grid of stat boxes: (number, label) tuples."""
    # 4 per row
    row1 = stats[:4]
    row2 = stats[4:]
    col_w = (PAGE_W - 2*MARGIN) / 4
    rows = []
    for row in [row1, row2]:
        cells = []
        for num, label in row:
            cell = Table([
                [Paragraph(str(num), style("sn", fontSize=20, textColor=GOLD,
                           fontName="Helvetica-Bold", leading=24, alignment=TA_CENTER))],
                [Paragraph(label, style("sl", fontSize=8, textColor=NAVY,
                           fontName="Helvetica-Bold", leading=10, alignment=TA_CENTER))],
            ], colWidths=[col_w - 4],
            style=TableStyle([
                ("BOX",(0,0),(-1,-1),1.5, NAVY),
                ("TOPPADDING",(0,0),(-1,-1),8),
                ("BOTTOMPADDING",(0,0),(-1,-1),8),
                ("LEFTPADDING",(0,0),(-1,-1),4),
                ("RIGHTPADDING",(0,0),(-1,-1),4),
                ("ALIGN",(0,0),(-1,-1),"CENTER"),
                ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
                ("BACKGROUND",(0,0),(-1,-1),WHITE),
            ]))
            cells.append(cell)
        grid_row = Table([cells], colWidths=[col_w]*4,
                         style=TableStyle([
                             ("ALIGN",(0,0),(-1,-1),"CENTER"),
                             ("TOPPADDING",(0,0),(-1,-1),2),
                             ("BOTTOMPADDING",(0,0),(-1,-1),2),
                             ("LEFTPADDING",(0,0),(-1,-1),2),
                             ("RIGHTPADDING",(0,0),(-1,-1),2),
                         ]))
        rows.append(grid_row)
        rows.append(Spacer(1,2*mm))
    return rows

def gold_insight_box(text, bold=False):
    fn = "Helvetica-Bold" if bold else "Helvetica"
    bg = GOLD if bold else WHITE
    tc = NAVY if bold else colors.HexColor("#333333")
    data = [[Paragraph(text, style("gi", fontSize=9.5, textColor=tc,
                       fontName=fn, leading=13, alignment=TA_CENTER))]]
    t = Table(data, colWidths=[PAGE_W - 2*MARGIN])
    t.setStyle(TableStyle([
        ("BOX",(0,0),(-1,-1),2, GOLD),
        ("BACKGROUND",(0,0),(-1,-1), bg),
        ("TOPPADDING",(0,0),(-1,-1),10),
        ("BOTTOMPADDING",(0,0),(-1,-1),10),
        ("LEFTPADDING",(0,0),(-1,-1),14),
        ("RIGHTPADDING",(0,0),(-1,-1),14),
    ]))
    return [t, Spacer(1,3*mm)]

def std_table(headers, rows, col_widths=None):
    """Standard table with navy header, alternating rows."""
    total_w = PAGE_W - 2*MARGIN
    if col_widths is None:
        col_widths = [total_w / len(headers)] * len(headers)
    data = [[Paragraph(f"<b>{h}</b>", style("th", fontSize=8.5, textColor=WHITE,
                       fontName="Helvetica-Bold", leading=11, alignment=TA_CENTER))
             for h in headers]]
    for i, row in enumerate(rows):
        bg = LGRAY if i % 2 == 0 else WHITE
        data.append([Paragraph(str(c), style("td", fontSize=8, textColor=NAVY,
                               fontName="Helvetica", leading=10, alignment=TA_CENTER))
                     for c in row])
    t = Table(data, colWidths=col_widths, repeatRows=1)
    style_cmds = [
        ("BACKGROUND",(0,0),(-1,0), NAVY),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[LGRAY, WHITE]),
        ("ALIGN",(0,0),(-1,-1),"CENTER"),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("TOPPADDING",(0,0),(-1,-1),4),
        ("BOTTOMPADDING",(0,0),(-1,-1),4),
        ("LEFTPADDING",(0,0),(-1,-1),4),
        ("RIGHTPADDING",(0,0),(-1,-1),4),
        ("GRID",(0,0),(-1,-1),0.5, colors.HexColor("#CCCCCC")),
        ("LINEBELOW",(0,0),(-1,0),1.5, GOLD),
    ]
    t.setStyle(TableStyle(style_cmds))
    return [t, Spacer(1,3*mm)]

def revenue_box(title, body):
    data = [
        [Paragraph(f"<b>{title}</b>", style("rv_t", fontSize=11, textColor=NAVY,
                   fontName="Helvetica-Bold", leading=13))],
        [Paragraph(body, style("rv_b", fontSize=9, textColor=colors.HexColor("#333333"),
                   fontName="Helvetica", leading=12))],
    ]
    t = Table(data, colWidths=[PAGE_W - 2*MARGIN])
    t.setStyle(TableStyle([
        ("BOX",(0,0),(-1,-1),1.5, NAVY),
        ("LINEABOVE",(0,0),(-1,0),3, GOLD),
        ("BACKGROUND",(0,0),(-1,-1), NAVY_LIGHT),
        ("TOPPADDING",(0,0),(-1,-1),8),
        ("BOTTOMPADDING",(0,0),(-1,-1),8),
        ("LEFTPADDING",(0,0),(-1,-1),10),
        ("RIGHTPADDING",(0,0),(-1,-1),10),
    ]))
    return [t, Spacer(1,3*mm)]

def market_box(title, body):
    data = [
        [Paragraph(f"<b>{title}</b>", style("mk_t", fontSize=12, textColor=GOLD,
                   fontName="Helvetica-Bold", leading=15))],
        [Paragraph(body, style("mk_b", fontSize=9.5, textColor=WHITE,
                   fontName="Helvetica", leading=13))],
    ]
    col_w = (PAGE_W - 2*MARGIN - 8) / 3
    t = Table(data, colWidths=[col_w])
    t.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,-1), NAVY),
        ("BOX",(0,0),(-1,-1),2, GOLD),
        ("TOPPADDING",(0,0),(-1,-1),10),
        ("BOTTOMPADDING",(0,0),(-1,-1),10),
        ("LEFTPADDING",(0,0),(-1,-1),10),
        ("RIGHTPADDING",(0,0),(-1,-1),10),
    ]))
    return t

def three_col_row(items):
    """Row of 3 equal boxes."""
    col_w = (PAGE_W - 2*MARGIN - 8) / 3
    cells = [items[0], Spacer(4,4), items[1], Spacer(4,4), items[2]]
    t = Table([[items[0], items[1], items[2]]],
              colWidths=[col_w, col_w, col_w],
              style=TableStyle([
                  ("ALIGN",(0,0),(-1,-1),"CENTER"),
                  ("VALIGN",(0,0),(-1,-1),"TOP"),
                  ("LEFTPADDING",(0,0),(-1,-1),3),
                  ("RIGHTPADDING",(0,0),(-1,-1),3),
              ]))
    return [t, Spacer(1,3*mm)]

def alloc_box(pct, label, color=NAVY):
    data = [
        [Paragraph(f"<b>{pct}</b>", style("al_p", fontSize=16, textColor=GOLD,
                   fontName="Helvetica-Bold", alignment=TA_CENTER))],
        [Paragraph(label, style("al_l", fontSize=9, textColor=WHITE,
                   fontName="Helvetica", leading=11, alignment=TA_CENTER))],
    ]
    col_w = (PAGE_W - 2*MARGIN - 12) / 4
    t = Table(data, colWidths=[col_w])
    t.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,-1), NAVY),
        ("BOX",(0,0),(-1,-1),1.5, GOLD),
        ("TOPPADDING",(0,0),(-1,-1),8),
        ("BOTTOMPADDING",(0,0),(-1,-1),8),
        ("LEFTPADDING",(0,0),(-1,-1),6),
        ("RIGHTPADDING",(0,0),(-1,-1),6),
        ("ALIGN",(0,0),(-1,-1),"CENTER"),
    ]))
    return t

def four_col_row(items):
    col_w = (PAGE_W - 2*MARGIN - 12) / 4
    t = Table([items], colWidths=[col_w]*4,
              style=TableStyle([
                  ("ALIGN",(0,0),(-1,-1),"CENTER"),
                  ("VALIGN",(0,0),(-1,-1),"TOP"),
                  ("LEFTPADDING",(0,0),(-1,-1),3),
                  ("RIGHTPADDING",(0,0),(-1,-1),3),
              ]))
    return [t, Spacer(1,3*mm)]

# ── watermark + footer canvas ─────────────────────────────────────────────────

def on_page(canvas, doc):
    canvas.saveState()
    # watermark
    canvas.setFont("Helvetica-Bold", 72)
    canvas.setFillColorRGB(0.93, 0.93, 0.93, alpha=0.25)
    canvas.translate(PAGE_W/2, PAGE_H/2)
    canvas.rotate(45)
    canvas.drawCentredString(0, 0, "CONFIDENTIAL")
    canvas.restoreState()

    canvas.saveState()
    # footer
    canvas.setFont("Helvetica", 7)
    canvas.setFillColor(colors.HexColor("#555555"))
    footer = "Confidential — Crystal Power Investments © 2026 | MatchPro Investor Package"
    canvas.drawCentredString(PAGE_W/2, 12*mm, footer)
    # page number
    canvas.drawRightString(PAGE_W - MARGIN, 12*mm, f"Page {doc.page}")
    # top rule
    canvas.setStrokeColor(GOLD)
    canvas.setLineWidth(1)
    canvas.line(MARGIN, PAGE_H - 14*mm, PAGE_W - MARGIN, PAGE_H - 14*mm)
    canvas.restoreState()

def on_cover(canvas, doc):
    """Cover page: full navy background, no footer rule."""
    canvas.saveState()
    canvas.setFillColor(NAVY)
    canvas.rect(0, 0, PAGE_W, PAGE_H, fill=1, stroke=0)
    canvas.restoreState()
    # watermark light on navy
    canvas.saveState()
    canvas.setFont("Helvetica-Bold", 72)
    canvas.setFillColorRGB(1, 1, 1, alpha=0.04)
    canvas.translate(PAGE_W/2, PAGE_H/2)
    canvas.rotate(45)
    canvas.drawCentredString(0, 0, "CONFIDENTIAL")
    canvas.restoreState()
    # footer
    canvas.saveState()
    canvas.setFont("Helvetica", 7)
    canvas.setFillColor(colors.HexColor("#AAAAAA"))
    canvas.drawCentredString(PAGE_W/2, 12*mm,
        "Confidential — Crystal Power Investments © 2026 | MatchPro Investor Package")
    canvas.drawRightString(PAGE_W - MARGIN, 12*mm, f"Page {doc.page}")
    canvas.restoreState()

# ── document build ─────────────────────────────────────────────────────────────

OUT_PATH = "/home/work/.openclaw/workspace/matchpro/MatchPro_Investor_Package_EN.pdf"

doc = SimpleDocTemplate(
    OUT_PATH,
    pagesize=A4,
    leftMargin=MARGIN,
    rightMargin=MARGIN,
    topMargin=20*mm,
    bottomMargin=20*mm,
    title="MatchPro Investor Package",
    author="Crystal Power Investments",
    subject="Investor Briefing",
)

story = []

# ══════════════════════════════════════════════════════════════════════════════
# PAGE 1 — COVER
# ══════════════════════════════════════════════════════════════════════════════
story.append(Spacer(1, 40*mm))

story.append(Paragraph("MatchPro™", style(
    "cov_title", fontSize=54, textColor=GOLD, fontName="Helvetica-Bold",
    alignment=TA_CENTER, leading=60)))

story.append(Spacer(1, 5*mm))

# Gold rule
rule_data = [[""]]
rule_t = Table(rule_data, colWidths=[PAGE_W - 2*MARGIN - 40*mm])
rule_t.setStyle(TableStyle([
    ("LINEABOVE",(0,0),(-1,-1),2,GOLD),
    ("LINEBELOW",(0,0),(-1,-1),2,GOLD),
    ("TOPPADDING",(0,0),(-1,-1),3),
    ("BOTTOMPADDING",(0,0),(-1,-1),3),
]))
story.append(Table([[rule_t]], colWidths=[PAGE_W - 2*MARGIN],
             style=TableStyle([("ALIGN",(0,0),(-1,-1),"CENTER")])))

story.append(Spacer(1, 10*mm))

story.append(Paragraph(
    "The Intelligence Layer for Egypt's Real Estate Market",
    style("cov_sub", fontSize=16, textColor=WHITE, fontName="Helvetica",
          alignment=TA_CENTER, leading=22)))

story.append(Spacer(1, 6*mm))

story.append(Paragraph(
    "Investor Briefing — Proof of Concept &amp; Commercial Case",
    style("cov_desc", fontSize=11, textColor=WHITE, fontName="Helvetica",
          alignment=TA_CENTER, leading=15)))

story.append(Spacer(1, 6*mm))

story.append(Paragraph(
    "May 2026 | Confidential — Not for Distribution",
    style("cov_date", fontSize=10, textColor=colors.HexColor("#CCCCCC"),
          fontName="Helvetica", alignment=TA_CENTER)))

story.append(Spacer(1, 8*mm))

story.append(Paragraph(
    "Crystal Power Investments | Cairo, Egypt",
    style("cov_co", fontSize=9, textColor=colors.HexColor("#BBBBBB"),
          fontName="Helvetica", alignment=TA_CENTER)))

story.append(PageBreak())

# ══════════════════════════════════════════════════════════════════════════════
# PAGE 2 — THE PROBLEM
# ══════════════════════════════════════════════════════════════════════════════
story.extend(navy_header("The Problem"))

story.extend(red_box(
    "Brokers Work in Silos",
    "50,000+ brokers in Greater Cairo operate across 200+ WhatsApp groups. A buyer in one group "
    "never sees a matching seller in another. Deals die because the right two people never meet."
))
story.extend(red_box(
    "No Real-Time Market Intelligence",
    "There is no live feed of buyer demand or seller supply in Egypt. Developers, investors, and "
    "brokers make decisions based on intuition, not data. Price discovery is broken."
))
story.extend(red_box(
    "Manual Matching Is Impossible at Scale",
    "A human broker can monitor 5-10 groups. MatchPro monitors 50+. The volume of daily messages "
    "(200-500/day) makes manual cross-referencing economically unviable."
))

story.append(PageBreak())

# ══════════════════════════════════════════════════════════════════════════════
# PAGE 3 — THE SOLUTION
# ══════════════════════════════════════════════════════════════════════════════
story.extend(navy_header("MatchPro — What It Does"))

story.append(Paragraph(
    "MatchPro is a real-time supply-demand matching engine built on top of Egypt's largest "
    "informal real estate data network — WhatsApp broker groups. It ingests, structures, and "
    "cross-matches every listing and buyer request across 50+ channels, surfacing actionable "
    "matches every 12 hours.",
    style("body", fontSize=10, textColor=colors.HexColor("#222222"), fontName="Helvetica",
          leading=15, spaceAfter=6)))
story.append(Spacer(1, 3*mm))

story.extend(navy_feature_box(
    "Live Market Ingestion",
    "Processes messages from 50+ broker WhatsApp channels continuously. Supply and demand signals "
    "are extracted, structured, and stored — creating a proprietary database that grows daily."
))
story.extend(navy_feature_box(
    "3-Dimension Matching Engine",
    "Every supply listing is scored against every demand request across Location, Price, and "
    "Specifications. Only matches scoring 70%+ are surfaced. Matches 90%+ are flagged as Very Hot Leads."
))
story.extend(navy_feature_box(
    "Intelligence Reports",
    "Every 12 hours, MatchPro generates ranked match reports for brokers, developers, and investors "
    "— segmented by area, transaction type, and confidence score."
))

story.append(PageBreak())

# ══════════════════════════════════════════════════════════════════════════════
# PAGE 4 — TRACTION
# ══════════════════════════════════════════════════════════════════════════════
story.extend(navy_header(
    "Real Traction — Real Data",
    "These are not projections. This is what MatchPro has already produced."
))

stats = [
    ("5,569",  "Broker Messages Processed"),
    ("4,000",  "Supply Listings Extracted"),
    ("7,626",  "Demand Requests Captured"),
    ("56,566", "AI Match Pairs Generated"),
    ("1,438",  "High-Confidence Matches (90%+)"),
    ("50+",    "WhatsApp Broker Channels"),
    ("40+",    "Cairo Areas Covered"),
    ("$97.5M", "Tracked Buyer Pipeline Value"),
]
story.extend(stat_grid(stats))

story.extend(gold_insight_box(
    "7,626 buyers are chasing 4,000 listings — a 91% demand surplus. MatchPro has mapped this "
    "entire gap in real time. No other platform in Egypt has done this."
))

story.append(PageBreak())

# ══════════════════════════════════════════════════════════════════════════════
# PAGE 5 — PROOF OF CONCEPT
# ══════════════════════════════════════════════════════════════════════════════
story.extend(navy_header(
    "The POC — One Cycle Covers The Commitment",
    "Based on the smallest 12-hour output: 27 Very Hot Leads (90%+ match score). Sample: 24 April 2026, 22:00."
))

poc_headers = ["KPI", "Conservative", "Base", "Optimistic"]
poc_rows = [
    ["Very Hot Leads per 12-hr cycle", "27", "27", "27"],
    ["Lead-to-close conversion", "2%", "5%", "8%"],
    ["Closed deals per cycle", "0.54", "1.35", "2.16"],
    ["Avg property price (USD)", "$120,000", "$257,000", "$450,000"],
    ["Commission rate", "2.0%", "2.5%", "2.5%"],
    ["Platform take-rate", "30%", "30%", "30%"],
    ["Net revenue per cycle (USD)", "$194", "$2,607", "$7,776"],
    ["Cycles per month (2/day)", "60", "60", "60"],
    ["Net monthly revenue (USD)", "$11,664", "$156,420", "$466,560"],
    ["Months to cover $250K", "21.4", "1.6", "0.5"],
]
col_ws = [(PAGE_W - 2*MARGIN)*0.42] + [(PAGE_W - 2*MARGIN)*0.19]*3
story.extend(std_table(poc_headers, poc_rows, col_ws))

story.extend(gold_insight_box(
    "Even in the most conservative scenario, the $250,000 commitment is recovered in 21 months "
    "from the SMALLEST possible output. Base case: 1.6 months. "
    "This is arithmetic applied to observed data.",
    bold=True
))

story.append(PageBreak())

# ══════════════════════════════════════════════════════════════════════════════
# PAGE 6 — BUSINESS MODEL
# ══════════════════════════════════════════════════════════════════════════════
story.extend(navy_header("How MatchPro Makes Money"))

story.extend(revenue_box(
    "Broker Subscription (SaaS)",
    "$50-$200/month per broker. TAM: 50,000+ brokers in Egypt. "
    "Target Year 1: 500 subscribers = $300K-$1.2M ARR"
))
story.extend(revenue_box(
    "Developer Intelligence Feed",
    "$500-$2,000/month per developer. Real-time demand data by area and unit type. "
    "Target Year 1: 20 developers = $120K-$480K ARR"
))
story.extend(revenue_box(
    "Verified Lead Sales",
    "$50-$200 per Very Hot Lead (90%+ match). 1,620 potential leads/month. "
    "Target Year 1 (10% monetized): $8K-$32K/month"
))
story.extend(revenue_box(
    "Transaction Commission Share",
    "30% of brokerage commission on closed deals. "
    "On a $250K property: $1,875 per closed deal"
))

story.append(PageBreak())

# ══════════════════════════════════════════════════════════════════════════════
# PAGE 7 — MARKET OPPORTUNITY
# ══════════════════════════════════════════════════════════════════════════════
story.extend(navy_header("The Market"))

m1 = market_box("TAM",
    "Egypt real estate: $97.5B+ annual pipeline. "
    "National market: EGP 1.2 trillion ($24B) annually.")
m2 = market_box("SAM",
    "Greater Cairo broker market: 50,000+ brokers. "
    "Annual commission pool: ~$600M. "
    "Target: 2% share = $12M ARR by Year 3.")
m3 = market_box("SOM",
    "Year 1: 500 broker subscribers + 20 developer feeds + lead sales "
    "= $1.5M-$2.5M revenue.")

story.extend(three_col_row([m1, m2, m3]))

story.append(Paragraph(
    "<b>Expansion:</b> The identical model applies to Saudi Arabia (Riyadh, Jeddah) and UAE "
    "(Dubai, Abu Dhabi) — markets with even higher transaction volumes. "
    "Crystal Power Investments already operates in Saudi Arabia.",
    style("exp", fontSize=9.5, textColor=colors.HexColor("#333333"), fontName="Helvetica",
          leading=13, spaceBefore=4)))

story.append(PageBreak())

# ══════════════════════════════════════════════════════════════════════════════
# PAGE 8 — COMPETITIVE ADVANTAGE
# ══════════════════════════════════════════════════════════════════════════════
story.extend(navy_header("Why MatchPro Wins"))

comp_headers = ["Feature", "MatchPro", "Aqarmap", "OLX", "Property Finder", "Manual Broker"]
comp_rows = [
    ["Real-time WhatsApp ingestion", "YES", "NO", "NO", "NO", "YES (1 group)"],
    ["Cross-group matching",          "YES", "NO", "NO", "NO", "NO"],
    ["50+ channel coverage",          "YES", "NO", "NO", "NO", "NO"],
    ["Demand surplus mapping",        "YES", "NO", "NO", "NO", "NO"],
    ["12-hour refresh cycle",         "YES", "NO", "NO", "NO", "NO"],
    ["3-dimension scoring",           "YES", "NO", "NO", "NO", "NO"],
    ["Informal market coverage",      "YES", "NO", "NO", "NO", "Partial"],
]
total_w = PAGE_W - 2*MARGIN
comp_cws = [total_w*0.30, total_w*0.14, total_w*0.14, total_w*0.10, total_w*0.18, total_w*0.14]
story.extend(std_table(comp_headers, comp_rows, comp_cws))

# Highlight YES in competitor table — done via post-processing style below
story.append(Paragraph(
    "<b>Structural Moat:</b> MatchPro's moat is the data network itself. Every broker channel "
    "added compounds the value for every other participant — a classic network effects business.",
    style("moat", fontSize=9.5, textColor=colors.HexColor("#333333"), fontName="Helvetica",
          leading=13, spaceBefore=4)))

story.append(PageBreak())

# ══════════════════════════════════════════════════════════════════════════════
# PAGE 9 — THE ASK
# ══════════════════════════════════════════════════════════════════════════════
story.extend(navy_header("The Investment Opportunity"))

ask_data = [[Paragraph(
    "Seeking: $250,000 – $2,000,000 USD",
    style("ask", fontSize=18, textColor=WHITE, fontName="Helvetica-Bold",
          alignment=TA_CENTER, leading=24))]]
ask_t = Table(ask_data, colWidths=[PAGE_W - 2*MARGIN])
ask_t.setStyle(TableStyle([
    ("BACKGROUND",(0,0),(-1,-1), NAVY),
    ("BOX",(0,0),(-1,-1),3, GOLD),
    ("TOPPADDING",(0,0),(-1,-1),14),
    ("BOTTOMPADDING",(0,0),(-1,-1),14),
]))
story.append(ask_t)
story.append(Spacer(1,4*mm))

a1 = alloc_box("40%", "Technology\nPlatform scaling, mobile app, API")
a2 = alloc_box("30%", "Sales & Growth\nBroker onboarding, developer partnerships")
a3 = alloc_box("20%", "Operations\nTeam, infrastructure, data quality")
a4 = alloc_box("10%", "Legal & IP\nPatents, licensing, regional expansion")
story.extend(four_col_row([a1, a2, a3, a4]))

mil_headers = ["Milestone", "Timeline", "KPI"]
mil_rows = [
    ["500 broker subscribers",  "Month 6",  "$25K MRR"],
    ["Developer product launch","Month 4",  "10 pilots"],
    ["Saudi Arabia pilot",      "Month 9",  "50 brokers"],
    ["$1M ARR",                 "Month 12", "Break-even"],
    ["Series A ready",          "Month 18", "$3M ARR"],
]
story.extend(std_table(mil_headers, mil_rows,
             [(PAGE_W-2*MARGIN)*0.50, (PAGE_W-2*MARGIN)*0.25, (PAGE_W-2*MARGIN)*0.25]))

story.extend(gold_insight_box(
    "Pre-money valuation: $1M – $6M USD. "
    "Comparable MENA PropTech at similar stage: $3M–$8M pre-money.",
    bold=True
))

story.append(PageBreak())

# ══════════════════════════════════════════════════════════════════════════════
# PAGE 10 — THE TEAM
# ══════════════════════════════════════════════════════════════════════════════
story.extend(navy_header("The Team"))

team_data = [
    [Paragraph("<b>Mo'men Maisara — Founder &amp; CEO</b>",
               style("t1", fontSize=13, textColor=NAVY, fontName="Helvetica-Bold", leading=17))],
    [Paragraph(
        "Multi-sector entrepreneur. BBA, Arab Academy. Operates across Egypt, Saudi Arabia, "
        "Germany, and Japan. Built MatchPro from zero to 56,566 AI matches and $97.5M tracked pipeline.",
        style("t2", fontSize=10, textColor=colors.HexColor("#333333"), fontName="Helvetica", leading=14))],
]
team_t = Table(team_data, colWidths=[PAGE_W - 2*MARGIN])
team_t.setStyle(TableStyle([
    ("BOX",(0,0),(-1,-1),1.5, NAVY),
    ("LINEABOVE",(0,0),(-1,0),4, GOLD),
    ("BACKGROUND",(0,0),(-1,-1), NAVY_LIGHT),
    ("TOPPADDING",(0,0),(-1,-1),10),
    ("BOTTOMPADDING",(0,0),(-1,-1),10),
    ("LEFTPADDING",(0,0),(-1,-1),14),
    ("RIGHTPADDING",(0,0),(-1,-1),14),
]))
story.append(team_t)
story.append(Spacer(1,4*mm))

cpi_data = [
    [Paragraph("<b>Crystal Power Investments</b>",
               style("c1", fontSize=13, textColor=NAVY, fontName="Helvetica-Bold", leading=17))],
    [Paragraph(
        "Founded February 2022. 366% ROI. 100+ projects. 98% client satisfaction. 4 countries. "
        "Active in real estate, schools catering (53 schools, 10,000+ students), hospitality, technology.",
        style("c2", fontSize=10, textColor=colors.HexColor("#333333"), fontName="Helvetica", leading=14))],
]
cpi_t = Table(cpi_data, colWidths=[PAGE_W - 2*MARGIN])
cpi_t.setStyle(TableStyle([
    ("BOX",(0,0),(-1,-1),1.5, NAVY),
    ("LINEABOVE",(0,0),(-1,0),4, GOLD),
    ("BACKGROUND",(0,0),(-1,-1), NAVY_LIGHT),
    ("TOPPADDING",(0,0),(-1,-1),10),
    ("BOTTOMPADDING",(0,0),(-1,-1),10),
    ("LEFTPADDING",(0,0),(-1,-1),14),
    ("RIGHTPADDING",(0,0),(-1,-1),14),
]))
story.append(cpi_t)

story.append(PageBreak())

# ══════════════════════════════════════════════════════════════════════════════
# PAGE 11 — RISK REGISTER
# ══════════════════════════════════════════════════════════════════════════════
story.extend(navy_header("Risk Register"))

risk_headers = ["Risk", "Likelihood", "Mitigation"]
risk_rows = [
    ["WhatsApp policy change",      "Medium", "Multi-channel ingestion + proprietary network"],
    ["Broker adoption resistance",  "Low",    "Free tier + demonstrated ROI before paywall"],
    ["Competitor replication",      "Low",    "12-month data moat + network effects"],
    ["Data privacy regulatory",     "Low",    "GDPR-aligned, no PII sold"],
    ["FX / Egypt macro risk",       "Medium", "USD pricing for enterprise"],
    ["Technical scaling",           "Low",    "Cloud-native, auto-scaling"],
]
story.extend(std_table(risk_headers, risk_rows,
             [(PAGE_W-2*MARGIN)*0.40, (PAGE_W-2*MARGIN)*0.15, (PAGE_W-2*MARGIN)*0.45]))

story.append(PageBreak())

# ══════════════════════════════════════════════════════════════════════════════
# PAGE 12 — NEXT STEPS
# ══════════════════════════════════════════════════════════════════════════════
story.extend(navy_header("Next Steps"))

ns_boxes = [
    ("1. NDA & Term Sheet (Week 1)",
     "Sign mutual NDA. Review draft term sheet. Confirm investment size and structure."),
    ("2. Due Diligence (Weeks 2-3)",
     "Full platform demo. Data room access. Reference calls with broker partners."),
    ("3. Close & Deploy (Week 4)",
     "Signed agreement. Wire transfer. First milestone review at 30 days."),
]
for title, body in ns_boxes:
    story.extend(navy_feature_box(title, body))

story.append(Spacer(1, 4*mm))

contact_data = [
    [Paragraph("<b>Mo'men Maisara — Founder &amp; CEO</b>",
               style("ct", fontSize=12, textColor=NAVY, fontName="Helvetica-Bold",
                     alignment=TA_CENTER))],
    [Paragraph(
        "maisaramoamen@outlook.com | +20 106 650 5665",
        style("ct2", fontSize=10, textColor=colors.HexColor("#333333"), fontName="Helvetica",
              alignment=TA_CENTER))],
    [Paragraph(
        "Crystal Power Investments | Cairo, Egypt",
        style("ct3", fontSize=10, textColor=colors.HexColor("#555555"), fontName="Helvetica",
              alignment=TA_CENTER))],
]
ct = Table(contact_data, colWidths=[PAGE_W - 2*MARGIN])
ct.setStyle(TableStyle([
    ("BOX",(0,0),(-1,-1),2, GOLD),
    ("BACKGROUND",(0,0),(-1,-1), GOLD_LIGHT),
    ("TOPPADDING",(0,0),(-1,-1),10),
    ("BOTTOMPADDING",(0,0),(-1,-1),10),
    ("LEFTPADDING",(0,0),(-1,-1),14),
    ("RIGHTPADDING",(0,0),(-1,-1),14),
]))
story.append(ct)

story.append(PageBreak())

# ══════════════════════════════════════════════════════════════════════════════
# PAGE 13 — DISCLAIMER
# ══════════════════════════════════════════════════════════════════════════════
story.extend(navy_header("Disclaimer"))

story.append(Spacer(1, 10*mm))
story.append(Paragraph(
    "This document has been prepared by Crystal Power Investments solely for informational "
    "purposes and does not constitute an offer to sell or a solicitation of an offer to buy "
    "any securities. The information is confidential and intended only for authorized recipients. "
    "Forward-looking statements involve known and unknown risks. Actual results may differ materially.",
    style("disc", fontSize=10, textColor=colors.HexColor("#444444"), fontName="Helvetica",
          leading=16, alignment=TA_JUSTIFY)))

# ── build PDF ─────────────────────────────────────────────────────────────────
def make_page_template(canvas, doc):
    if doc.page == 1:
        on_cover(canvas, doc)
    else:
        on_page(canvas, doc)

doc.build(story, onFirstPage=make_page_template, onLaterPages=make_page_template)
print(f"PDF saved: {OUT_PATH}")


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL — Top 100 Matches
# ══════════════════════════════════════════════════════════════════════════════
print("Generating Excel...")

XLSX_IN  = "/home/work/.openclaw/media/inbound/CrystalPower-MatchPro-2026-05-07---520c25a7-c206-4c41-a3a1-f3c168564b6d.xlsx"
XLSX_OUT = "/home/work/.openclaw/workspace/matchpro/MatchPro_Top100_Clean_EN.xlsx"

wb_in = openpyxl.load_workbook(XLSX_IN)
ws_in = wb_in.active

# Map headers (row 4)
header_map = {}
for col in range(1, ws_in.max_column + 1):
    val = ws_in.cell(4, col).value
    if val:
        header_map[val] = col

def col(name):
    return header_map.get(name, None)

# Read all data rows (row 5+)
rows = []
for r in range(5, ws_in.max_row + 1):
    seller_phone = ws_in.cell(r, col("Seller Phone")).value
    buyer_phone  = ws_in.cell(r, col("Buyer Phone")).value
    # skip same-person matches
    if seller_phone and buyer_phone and str(seller_phone).strip() == str(buyer_phone).strip():
        continue
    score_raw = ws_in.cell(r, col("Score %")).value
    try:
        score = float(score_raw) if score_raw is not None else 0.0
    except (ValueError, TypeError):
        score = 0.0
    if score == 0.0 and score_raw is None:
        continue  # skip blank rows
    rows.append({
        "score": score,
        "transaction": ws_in.cell(r, col("Transaction")).value,
        "property_type": ws_in.cell(r, col("Property Type")).value,
        "location": ws_in.cell(r, col("Location")).value,
        "price_egp": ws_in.cell(r, col("Price (EGP)")).value,
        "seller_name": ws_in.cell(r, col("Seller Name")).value,
        "seller_phone": seller_phone,
        "buyer_name": ws_in.cell(r, col("Buyer Name")).value,
        "buyer_phone": buyer_phone,
        "budget_egp": ws_in.cell(r, col("Budget Max (EGP)")).value,
        "loc_score": ws_in.cell(r, col("Location Score %")).value,
        "price_score": ws_in.cell(r, col("Price Score %")).value,
        "specs_score": ws_in.cell(r, col("Specs Score %")).value,
        "match_date": ws_in.cell(r, col("Match Date")).value,
    })

# Sort and take top 100
rows.sort(key=lambda x: x["score"], reverse=True)
top100 = rows[:100]

# Build output workbook
wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = "Top 100 Matches"

NAVY_HEX  = "0B2545"
WHITE_HEX = "FFFFFF"
LBLUE_HEX = "F0F4FF"
GOLD_HEX  = "C9A961"

navy_fill  = PatternFill("solid", fgColor=NAVY_HEX)
white_fill = PatternFill("solid", fgColor=WHITE_HEX)
blue_fill  = PatternFill("solid", fgColor=LBLUE_HEX)
gold_fill  = PatternFill("solid", fgColor=GOLD_HEX)

header_font = Font(name="Calibri", bold=True, color=WHITE_HEX, size=11)
data_font   = Font(name="Calibri", size=10)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=False)

thin = Side(style="thin", color="CCCCCC")
thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

out_headers = [
    "Rank", "Score%", "Transaction", "Property Type", "Location",
    "Price EGP", "Seller Name", "Seller Phone", "Buyer Name", "Buyer Phone",
    "Buyer Budget EGP", "Location Score%", "Price Score%", "Specs Score%", "Match Date"
]

ws_out.append(out_headers)

# Style header row
for c, _ in enumerate(out_headers, 1):
    cell = ws_out.cell(1, c)
    cell.fill = navy_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = thin_border

# Data rows
for i, row in enumerate(top100, 1):
    excel_row = i + 1
    fill = white_fill if i % 2 == 1 else blue_fill

    vals = [
        i,
        row["score"],
        row["transaction"],
        row["property_type"],
        row["location"],
        row["price_egp"],
        row["seller_name"],
        row["seller_phone"],
        row["buyer_name"],
        row["buyer_phone"],
        row["budget_egp"],
        row["loc_score"],
        row["price_score"],
        row["specs_score"],
        row["match_date"],
    ]
    ws_out.append(vals)

    for c, _ in enumerate(vals, 1):
        cell = ws_out.cell(excel_row, c)
        cell.border = thin_border
        cell.alignment = center_align
        # Gold fill for Score >= 90
        if c == 2:  # Score% column
            try:
                if float(row["score"]) >= 90:
                    cell.fill = gold_fill
                else:
                    cell.fill = fill
            except:
                cell.fill = fill
        else:
            cell.fill = fill
        cell.font = data_font

# Freeze top row
ws_out.freeze_panes = "A2"

# Auto-width columns
col_widths = {
    "A": 6,  # Rank
    "B": 10, # Score%
    "C": 14, # Transaction
    "D": 16, # Property Type
    "E": 20, # Location
    "F": 16, # Price EGP
    "G": 20, # Seller Name
    "H": 16, # Seller Phone
    "I": 20, # Buyer Name
    "J": 16, # Buyer Phone
    "K": 18, # Buyer Budget EGP
    "L": 16, # Location Score%
    "M": 14, # Price Score%
    "N": 14, # Specs Score%
    "O": 16, # Match Date
}
for col_letter, width in col_widths.items():
    ws_out.column_dimensions[col_letter].width = width

# Row height
ws_out.row_dimensions[1].height = 22
for r in range(2, 102):
    ws_out.row_dimensions[r].height = 18

wb_out.save(XLSX_OUT)
print(f"Excel saved: {XLSX_OUT}")
print(f"Top 100 records written (from {len(rows)} valid matches).")
