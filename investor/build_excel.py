#!/usr/bin/env python3
"""Build MatchPro™ Investor Summary Excel"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "MatchPro_Investment_Case"

# ── Color palette ──────────────────────────────────────────
NAVY   = "1F3864"
GOLD   = "C9A84C"
WHITE  = "FFFFFF"
ICE    = "E8F4FD"
MUTED  = "8FAABF"
DARK   = "0A1628"
GREEN  = "2ECC71"
LIGHT_NAVY = "243F60"
MID_NAVY   = "162D4A"

# ── Helper fills ───────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def border(style="thin", color="1E3A5F"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def thick_border(color="C9A84C"):
    t = Side(style="medium", color=color)
    return Border(left=t, right=t, top=t, bottom=t)

def font(bold=False, color=WHITE, size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Calibri")

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# ── Column widths ──────────────────────────────────────────
col_widths = {
    1: 32, 2: 18, 3: 18, 4: 16, 5: 20, 6: 18, 7: 16
}
for col, w in col_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w

# ═══════════════════════════════════════════════════════════
# SECTION A — PLATFORM STATS (rows 1–15)
# ═══════════════════════════════════════════════════════════

def merge_style(ws, cell_range, text, bg, fg=WHITE, sz=12, bold=True,
                h_align="center", wrap=False):
    ws.merge_cells(cell_range)
    top_left = ws[cell_range.split(":")[0]]
    top_left.value = text
    top_left.font = font(bold=bold, color=fg, size=sz)
    top_left.fill = fill(bg)
    top_left.alignment = align(h_align, "center", wrap)
    top_left.border = border()

# Row 1 – Main Title
merge_style(ws, "A1:G1",
    "MatchPro™ — Real Estate Intelligence Platform",
    NAVY, GOLD, sz=16, bold=True)
ws.row_dimensions[1].height = 28

# Row 2 – Subtitle
merge_style(ws, "A2:G2",
    "Investment Summary  |  May 2026  |  Crystal Power Investments  |  CEO: Mo'men Maisara",
    LIGHT_NAVY, ICE, sz=10, bold=False)
ws.row_dimensions[2].height = 18

# Row 3 – blank separator
ws.row_dimensions[3].height = 6

# Row 4 – Section header
merge_style(ws, "A4:G4", "KEY PERFORMANCE INDICATORS", MID_NAVY, GOLD, sz=11)
ws.row_dimensions[4].height = 20

# Rows 5–6: KPI headers
kpi_headers = ["Demand Requests", "Supply Listings", "AI Matches Generated",
               "High-Conf. Leads ≥90%", "Areas — Greater Cairo", "Pipeline Value (USD)"]
kpi_values  = ["7,626", "4,000+", "56,566", "1,438", "40+", "$97.5M"]

for i, (h, v) in enumerate(zip(kpi_headers, kpi_values)):
    col = i + 1
    if col > 6:
        break
    cell_h = ws.cell(row=5, column=col)
    cell_h.value = h
    cell_h.font = font(bold=True, color=ICE, size=9)
    cell_h.fill = fill(NAVY)
    cell_h.alignment = align("center", "center", wrap=True)
    cell_h.border = border()

    cell_v = ws.cell(row=6, column=col)
    cell_v.value = v
    cell_v.font = font(bold=True, color=GOLD, size=16)
    cell_v.fill = fill(MID_NAVY)
    cell_v.alignment = align("center", "center")
    cell_v.border = border("medium", GOLD)

ws.row_dimensions[5].height = 30
ws.row_dimensions[6].height = 36

# Row 7 – EGP Pipeline note
merge_style(ws, "A7:G7",
    "EGP 4.87 Billion in verified buyer budgets  |  50+ broker WhatsApp channels monitored  |  27 high-confidence matches per cycle",
    DARK, GOLD, sz=9, bold=False, h_align="center")
ws.row_dimensions[7].height = 16

# Rows 8–15 blank / separator
for r in range(8, 16):
    ws.row_dimensions[r].height = 4

# ═══════════════════════════════════════════════════════════
# SECTION B — THE NUMBERS (rows 17–35)
# ═══════════════════════════════════════════════════════════

merge_style(ws, "A17:G17", "INVESTMENT ROI — THREE-SCENARIO MODEL", MID_NAVY, GOLD, sz=11)
ws.row_dimensions[17].height = 20

# Table headers
b_headers = ["Scenario", "Monthly Revenue", "Recovery of $250K", "Deals/Month",
             "Commission Base", "Platform Take (30%)", "Net Revenue"]
for c, h in enumerate(b_headers, 1):
    cell = ws.cell(row=18, column=c)
    cell.value = h
    cell.font = font(bold=True, color=GOLD, size=9)
    cell.fill = fill(NAVY)
    cell.alignment = align("center", "center", wrap=True)
    cell.border = border()
ws.row_dimensions[18].height = 28

# Data rows
b_data = [
    ("Conservative", "$23,328", "10.7 months", "2–3", "$2,333/deal", "~$700", "~$23K/mo"),
    ("⭐  Base Case", "$156,127", "1.6 months",  "8–12", "$10,000/deal", "$3,000", "~$156K/mo"),
    ("Optimistic",   "$437,400", "17 days",      "20+", "$18,000/deal", "$5,400", "~$437K/mo"),
]
b_row_colors = [MID_NAVY, "1A3050", MID_NAVY]
b_val_colors = [GOLD, "E2C472", GREEN]

for i, (row_data, bg, vc) in enumerate(zip(b_data, b_row_colors, b_val_colors)):
    r = 19 + i
    for c, val in enumerate(row_data, 1):
        cell = ws.cell(row=r, column=c)
        cell.value = val
        cell.fill = fill(bg)
        cell.border = border()
        cell.alignment = align("center", "center")
        if c == 1:
            cell.font = font(bold=True, color=WHITE if i != 1 else GOLD, size=10)
        else:
            cell.font = font(bold=(c == 2), color=vc if c == 2 else ICE, size=10)
    ws.row_dimensions[r].height = 22

# Unit Economics header
merge_style(ws, "A23:G23", "UNIT ECONOMICS", MID_NAVY, GOLD, sz=11)
ws.row_dimensions[23].height = 20

ue_data = [
    ("Average Deal Size",      "$116,000 – $450,000",   "Madinaty apartment to New Cairo villa"),
    ("Commission Rate",        "2 – 2.5%",               "Egypt market standard"),
    ("Platform Take-Rate",     "30%",                    "Of generated commission"),
    ("Net Per Deal",           "$720 – $3,375",          "Platform net revenue per closed deal"),
    ("High-Confidence Leads",  "1,438 active / 27 per cycle", "Confidence score ≥90%"),
    ("Cycle Frequency",        "Every 12 hours",         "Fully automated — zero broker input"),
]
for i, (label, val, note) in enumerate(ue_data):
    r = 24 + i
    bg = NAVY if i % 2 == 0 else MID_NAVY

    cell_l = ws.cell(row=r, column=1)
    cell_l.value = label
    cell_l.font = font(bold=True, color=ICE, size=9)
    cell_l.fill = fill(bg)
    cell_l.border = border()
    cell_l.alignment = align("left", "center")

    cell_v = ws.cell(row=r, column=2)
    cell_v.value = val
    cell_v.font = font(bold=True, color=GOLD, size=10)
    cell_v.fill = fill(bg)
    cell_v.border = border()
    cell_v.alignment = align("center", "center")

    # Merge columns 3-7 for note
    ws.merge_cells(f"C{r}:G{r}")
    cell_n = ws.cell(row=r, column=3)
    cell_n.value = note
    cell_n.font = font(bold=False, color=MUTED, size=9)
    cell_n.fill = fill(bg)
    cell_n.border = border()
    cell_n.alignment = align("left", "center")
    ws.row_dimensions[r].height = 18

# Row 30 separator
for r in range(30, 36):
    ws.row_dimensions[r].height = 4

# ═══════════════════════════════════════════════════════════
# SECTION C — MARKET DATA (rows 37–55)
# ═══════════════════════════════════════════════════════════

merge_style(ws, "A37:G37", "GREATER CAIRO MARKET COVERAGE — AREA BREAKDOWN", MID_NAVY, GOLD, sz=11)
ws.row_dimensions[37].height = 20

c_headers = ["Area", "Demand Requests", "Supply Listings", "D/S Ratio",
             "Avg Budget (EGP)", "Zone Status", "Notes"]
for c, h in enumerate(c_headers, 1):
    cell = ws.cell(row=38, column=c)
    cell.value = h
    cell.font = font(bold=True, color=GOLD, size=9)
    cell.fill = fill(NAVY)
    cell.alignment = align("center", "center", wrap=True)
    cell.border = border()
ws.row_dimensions[38].height = 25

c_data = [
    ("Madinaty",     "527",   "4,000 (platform)", "1.91×", "5.8M",  "🔥 Hottest",  "Largest supply base; agent HQ"),
    ("New Cairo",    "197",   "—",                 "—",     "12.24M","🟢 Premium",  "High-value transactions"),
    ("Al-Rehab",     "69",    "—",                 "—",     "4.75M", "🟢 Active",   "Steady demand volume"),
    ("6th October",  "54",    "—",                 "—",     "8M",    "🟡 Growing",  "Emerging buyer interest"),
    ("Sheikh Zayed", "27",    "—",                 "—",     "22M",   "💎 Luxury",   "Ultra-high-net-worth segment"),
    ("Other Zones",  "508",   "—",                 "—",     "—",     "📍 Coverage", "35+ additional areas tracked"),
]
for i, row_data in enumerate(c_data):
    r = 39 + i
    bg = MID_NAVY if i % 2 == 0 else NAVY
    for c, val in enumerate(row_data, 1):
        cell = ws.cell(row=r, column=c)
        cell.value = val
        cell.fill = fill(bg)
        cell.border = border()
        cell.alignment = align("center" if c > 1 else "left", "center")
        if c == 1:
            cell.font = font(bold=True, color=WHITE, size=9)
        elif c == 6:
            cell.font = font(bold=True, color=GOLD, size=9)
        else:
            cell.font = font(bold=False, color=ICE, size=9)
    ws.row_dimensions[r].height = 18

# Total row
r = 45
merge_cells_list = []
total_data = ["TOTAL", "1,382", "4,000+", "1.91×", "$97.5M Pipeline (USD)", "", ""]
for c, val in enumerate(total_data, 1):
    cell = ws.cell(row=r, column=c)
    cell.value = val
    cell.font = font(bold=True, color=NAVY, size=10)
    cell.fill = fill(GOLD)
    cell.border = border("medium", GOLD)
    cell.alignment = align("center" if c > 1 else "left", "center")
ws.merge_cells("E45:G45")
ws.cell(row=45, column=5).value = "$97.5M Pipeline (USD)"
ws.cell(row=45, column=5).font = font(bold=True, color=NAVY, size=10)
ws.cell(row=45, column=5).fill = fill(GOLD)
ws.cell(row=45, column=5).alignment = align("center", "center")
ws.row_dimensions[r].height = 22

# Platform note
merge_style(ws, "A46:G46",
    "Platform monitors 50+ WhatsApp channels | Arabic + English NLP | 40+ areas tracked | Auto-refresh every 12 hours",
    DARK, MUTED, sz=9, bold=False, h_align="center")
ws.row_dimensions[46].height = 16

# Rows 47-56 separator
for r in range(47, 57):
    ws.row_dimensions[r].height = 4

# ═══════════════════════════════════════════════════════════
# SECTION D — INVESTMENT ASK (rows 57–70)
# ═══════════════════════════════════════════════════════════

merge_style(ws, "A57:G57", "INVESTMENT OPPORTUNITY — SEED ROUND", MID_NAVY, GOLD, sz=11)
ws.row_dimensions[57].height = 20

# Ask box
merge_style(ws, "A58:G58",
    "Seeking: $100,000 – $500,000  |  Stage: Seed  |  Operator: Mo'men Maisara, Crystal Power Investments",
    NAVY, GOLD, sz=11, bold=True, h_align="center")
ws.row_dimensions[58].height = 22

# Use of Funds header
d_headers = ["Use of Funds", "Allocation %", "Key Hires / Actions", "Target Outcome", "", "", ""]
for c, h in enumerate(d_headers[:4], 1):
    cell = ws.cell(row=59, column=c)
    cell.value = h
    cell.font = font(bold=True, color=GOLD, size=9)
    cell.fill = fill(NAVY)
    cell.alignment = align("center", "center", wrap=True)
    cell.border = border()
ws.merge_cells("E59:G59")
ws.row_dimensions[59].height = 22

fund_data = [
    ("⚙️  Engineering Team",    "50%", "3 senior engineers",     "Platform scale-out & API integrations"),
    ("📣  Sales Expansion",     "30%", "2 sales reps + marketing","Broker acquisition across 5 cities"),
    ("🔑  Platform Licensing",  "20%", "Legal + onboarding",      "SaaS contracts, broker portal launch"),
]
for i, (use, pct, actions, outcome) in enumerate(fund_data):
    r = 60 + i
    bg = MID_NAVY if i % 2 == 0 else NAVY
    for c, val in enumerate([use, pct, actions, outcome], 1):
        cell = ws.cell(row=r, column=c)
        cell.value = val
        cell.fill = fill(bg)
        cell.border = border()
        cell.alignment = align("left" if c in (1, 3, 4) else "center", "center")
        if c == 1:
            cell.font = font(bold=True, color=ICE, size=9)
        elif c == 2:
            cell.font = font(bold=True, color=GOLD, size=11)
        else:
            cell.font = font(bold=False, color=ICE, size=9)
    ws.merge_cells(f"D{r}:G{r}")
    ws.row_dimensions[r].height = 20

# 12-Month Milestones
merge_style(ws, "A63:G63", "12-MONTH MILESTONES", MID_NAVY, GOLD, sz=11)
ws.row_dimensions[63].height = 20

milestones = [
    ("Month 3",  "30 broker licenses",   "$15K MRR",    "Platform beta launch — Madinaty + New Cairo"),
    ("Month 6",  "80 broker licenses",   "$40K MRR",    "Expand to 6th October, Sheikh Zayed"),
    ("Month 9",  "140 broker licenses",  "$70K MRR",    "Arabic NLP v2 — full dialect support"),
    ("Month 12", "200 broker licenses",  "$100K MRR 🎯","National rollout — Alexandria + Hurghada"),
]
mile_headers = ["Timeline", "Broker Licenses", "Target MRR", "Key Milestone"]
for c, h in enumerate(mile_headers, 1):
    cell = ws.cell(row=64, column=c)
    cell.value = h
    cell.font = font(bold=True, color=GOLD, size=9)
    cell.fill = fill(NAVY)
    cell.alignment = align("center", "center")
    cell.border = border()
ws.merge_cells("D64:G64")
ws.row_dimensions[64].height = 20

for i, (mo, lic, mrr, key) in enumerate(milestones):
    r = 65 + i
    bg = MID_NAVY if i % 2 == 0 else NAVY
    vc = GOLD if i < 3 else GREEN
    for c, val in enumerate([mo, lic, mrr, key], 1):
        cell = ws.cell(row=r, column=c)
        cell.value = val
        cell.fill = fill(bg)
        cell.border = border()
        cell.alignment = align("center" if c < 4 else "left", "center")
        cell.font = font(bold=(c == 3), color=vc if c == 3 else (GOLD if c == 1 else ICE), size=9)
    ws.merge_cells(f"D{r}:G{r}")
    ws.row_dimensions[r].height = 18

# Final tagline
merge_style(ws, "A69:G69",
    '"The market is talking — are you listening?  |  MatchPro™ is a trademark of Crystal Power Investments  |  mmaisara@crystalpowerinvestment.com"',
    DARK, GOLD, sz=9, bold=False, h_align="center", wrap=True)
ws.row_dimensions[69].height = 20

# ── Save ───────────────────────────────────────────────────
out = "/home/work/.openclaw/workspace/matchpro/investor/MatchPro_Investor_Summary_2026.xlsx"
wb.save(out)
print(f"Saved: {out}")
import os
size = os.path.getsize(out)
print(f"Size: {size:,} bytes ({size/1024:.1f} KB)")
